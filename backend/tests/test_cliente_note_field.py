"""Test that the `note` field on a cliente is:
- returned by GET /api/clienti/{id}
- updatable via PUT /api/clienti/{id}
- exported in the Excel export at column BM (65th column)
"""
import os
import time
import pytest
import requests
from io import BytesIO
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL').rstrip('/')
CLIENTE_ID = "3818627c-eb7a-4421-8243-5d9be37f552f"


@pytest.fixture(scope="module")
def auth_token():
    resp = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"username": "admin", "password": "admin123"},
        timeout=30,
    )
    assert resp.status_code == 200, f"Login failed: {resp.status_code} {resp.text}"
    data = resp.json()
    token = data.get("access_token") or data.get("token")
    assert token, f"No token in response: {data}"
    return token


@pytest.fixture(scope="module")
def headers(auth_token):
    return {"Authorization": f"Bearer {auth_token}", "Content-Type": "application/json"}


# ---- GET / PUT persistence for `note` ----

def test_get_cliente_has_note_field(headers):
    resp = requests.get(f"{BASE_URL}/api/clienti/{CLIENTE_ID}", headers=headers, timeout=30)
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert "note" in data, "note field missing from cliente response"
    # Preserve original note for later restore
    assert data["note"], f"Expected note to be populated, got: {data['note']!r}"


def test_put_cliente_updates_note_and_persists(headers):
    # Read original note
    orig = requests.get(f"{BASE_URL}/api/clienti/{CLIENTE_ID}", headers=headers, timeout=30).json()
    original_note = orig.get("note") or ""

    new_note = f"TEST_note_update_{int(time.time())}"
    # PUT update
    resp = requests.put(
        f"{BASE_URL}/api/clienti/{CLIENTE_ID}",
        headers=headers,
        json={"note": new_note},
        timeout=30,
    )
    assert resp.status_code in (200, 204), f"PUT failed: {resp.status_code} {resp.text}"

    # GET to verify persistence
    verify = requests.get(f"{BASE_URL}/api/clienti/{CLIENTE_ID}", headers=headers, timeout=30)
    assert verify.status_code == 200
    assert verify.json().get("note") == new_note, f"Note not persisted, got: {verify.json().get('note')!r}"

    # Restore original note
    restore = requests.put(
        f"{BASE_URL}/api/clienti/{CLIENTE_ID}",
        headers=headers,
        json={"note": original_note},
        timeout=30,
    )
    assert restore.status_code in (200, 204)
    final = requests.get(f"{BASE_URL}/api/clienti/{CLIENTE_ID}", headers=headers, timeout=30).json()
    assert final.get("note") == original_note


# ---- Excel export: column BM (65th) should be "Note" and contain `note` value ----

def test_excel_export_column_bm_contains_note(headers):
    # Auth via Authorization header for the export
    resp = requests.get(
        f"{BASE_URL}/api/clienti/export/excel",
        headers={"Authorization": headers["Authorization"]},
        timeout=90,
    )
    assert resp.status_code == 200, f"Excel export failed: {resp.status_code} {resp.text[:300]}"
    ct = resp.headers.get("content-type", "")
    assert "spreadsheet" in ct or "excel" in ct or "octet-stream" in ct, f"Unexpected content-type: {ct}"

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active

    # Column BM = 65th column
    bm_header = ws.cell(row=1, column=65).value
    assert bm_header and "Note" in str(bm_header), f"BM header not 'Note', got: {bm_header!r}"

    # Find our test cliente row and check note value
    # First find the "id" or code column: iterate rows and search for CLIENTE_ID or use first column heuristics.
    # We will search all rows for the note value we know: it starts with "Test preservazione tipologia energia_fastweb_tls"
    found = False
    expected_prefix = "Test preservazione tipologia energia_fastweb_tls"
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=65).value
        if val and expected_prefix in str(val):
            found = True
            break
    assert found, f"Expected note starting with {expected_prefix!r} not found in column BM of export"
