"""Helper condivisi: province, assegnazione lead, export Excel, import clienti
(estratti da server.py - refactoring fase 3)."""
import asyncio
import io
import json
import logging
import re
import uuid
from datetime import datetime, timezone, timedelta, date, time
from typing import List, Optional, Dict, Any, Tuple

try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except ImportError:  # pragma: no cover
    from backports.zoneinfo import ZoneInfo  # type: ignore

# NEW (feb 2026): fuso orario di riferimento dell'app (Europa/Roma, gestisce CET/CEST)
APP_TIMEZONE = ZoneInfo("Europe/Rome")


def rome_date_to_utc_range(date_str: str, tz_name: Optional[str] = None) -> Tuple[datetime, datetime]:
    """Converte una data local (YYYY-MM-DD) nell'intervallo UTC corrispondente
    `[00:00, 23:59:59.999999]` nel fuso `tz_name` (default Europe/Rome).

    Esempio (estate, +02:00): "2026-08-15" → (2026-08-14T22:00:00Z, 2026-08-15T21:59:59.999999Z).
    Gestisce automaticamente CET/CEST tramite zoneinfo. Se `tz_name` è valorizzato
    (es. fuso scelto dall'utente), usa quel fuso al posto di Europe/Rome.
    """
    tz = APP_TIMEZONE
    if tz_name:
        try:
            tz = ZoneInfo(tz_name)
        except Exception:
            tz = APP_TIMEZONE
    d = datetime.fromisoformat(date_str).date()
    start_rome = datetime.combine(d, time(0, 0, 0, 0), tzinfo=tz)
    end_rome = datetime.combine(d, time(23, 59, 59, 999999), tzinfo=tz)
    return start_rome.astimezone(timezone.utc), end_rome.astimezone(timezone.utc)


import pandas as pd
import openpyxl
import tempfile
from openpyxl.styles import Font, PatternFill, Alignment
from notifications import notify_agent_new_lead
from fastapi import HTTPException, UploadFile

from database import db
from models import *  # noqa: F401,F403

# Italian Provinces (111 provinces)
ITALIAN_PROVINCES = [
    "Agrigento", "Alessandria", "Ancona", "Aosta", "Arezzo", "Ascoli Piceno", "Asti", "Avellino", "Bari", 
    "Barletta-Andria-Trani", "Belluno", "Benevento", "Bergamo", "Biella", "Bologna", "Bolzano", 
    "Brescia", "Brindisi", "Cagliari", "Caltanissetta", "Campobasso", "Caserta", 
    "Catania", "Catanzaro", "Chieti", "Como", "Cosenza", "Cremona", "Crotone", "Cuneo", "Enna", 
    "Fermo", "Ferrara", "Firenze", "Foggia", "Forlì-Cesena", "Frosinone", "Gallura Nord-Est Sardegna", "Genova", "Gorizia", 
    "Grosseto", "Imperia", "Isernia", "L'Aquila", "La Spezia", "Latina", "Lecce", "Lecco", "Livorno", 
    "Lodi", "Lucca", "Macerata", "Mantova", "Massa-Carrara", "Matera", "Messina", "Milano", "Modena", 
    "Monza e Brianza", "Napoli", "Novara", "Nuoro", "Oristano", "Padova", "Palermo", 
    "Parma", "Pavia", "Perugia", "Pesaro e Urbino", "Pescara", "Piacenza", "Pisa", "Pistoia", "Pordenone", 
    "Potenza", "Prato", "Ragusa", "Ravenna", "Reggio Calabria", "Reggio Emilia", "Rieti", "Rimini", 
    "Roma", "Rovigo", "Salerno", "Medio Campidano", "Sassari", "Savona", "Siena", "Siracusa", "Sondrio", 
    "Sulcis Iglesiente", "Taranto", "Teramo", "Terni", "Torino", "Ogliastra", "Trapani", "Trento", "Treviso", "Trieste", 
    "Udine", "Varese", "Venezia", "Verbano-Cusio-Ossola", "Vercelli", "Verona", "Vibo Valentia", 
    "Vicenza", "Viterbo"
]

# Province mapping: Nome completo → Sigla (per Zapier che invia nomi completi)
PROVINCE_TO_CODE = {
    "Agrigento": "AG", "Alessandria": "AL", "Ancona": "AN", "Aosta": "AO", "Arezzo": "AR",
    "Ascoli Piceno": "AP", "Asti": "AT", "Avellino": "AV", "Bari": "BA", "Barletta-Andria-Trani": "BT",
    "Belluno": "BL", "Benevento": "BN", "Bergamo": "BG", "Biella": "BI", "Bologna": "BO",
    "Bolzano": "BZ", "Brescia": "BS", "Brindisi": "BR", "Cagliari": "CA", "Caltanissetta": "CL",
    "Campobasso": "CB", "Caserta": "CE", "Catania": "CT", "Catanzaro": "CZ", "Chieti": "CH",
    "Como": "CO", "Cosenza": "CS", "Cremona": "CR", "Crotone": "KR", "Cuneo": "CN",
    "Enna": "EN", "Fermo": "FM", "Ferrara": "FE", "Firenze": "FI", "Foggia": "FG",
    "Forlì-Cesena": "FC", "Frosinone": "FR", "Genova": "GE", "Gorizia": "GO", "Grosseto": "GR",
    "Imperia": "IM", "Isernia": "IS", "L'Aquila": "AQ", "La Spezia": "SP", "Latina": "LT",
    "Lecce": "LE", "Lecco": "LC", "Livorno": "LI", "Lodi": "LO", "Lucca": "LU",
    "Macerata": "MC", "Mantova": "MN", "Massa-Carrara": "MS", "Matera": "MT", "Messina": "ME",
    "Milano": "MI", "Modena": "MO", "Monza e Brianza": "MB", "Napoli": "NA", "Novara": "NO",
    "Nuoro": "NU", "Oristano": "OR", "Padova": "PD", "Palermo": "PA", "Parma": "PR",
    "Pavia": "PV", "Perugia": "PG", "Pesaro e Urbino": "PU", "Pescara": "PE", "Piacenza": "PC",
    "Pisa": "PI", "Pistoia": "PT", "Pordenone": "PN", "Potenza": "PZ", "Prato": "PO",
    "Ragusa": "RG", "Ravenna": "RA", "Reggio Calabria": "RC", "Reggio Emilia": "RE", "Rieti": "RI",
    "Rimini": "RN", "Roma": "RM", "Rovigo": "RO", "Salerno": "SA", "Sassari": "SS",
    "Savona": "SV", "Siena": "SI", "Siracusa": "SR", "Sondrio": "SO", "Taranto": "TA",
    "Teramo": "TE", "Terni": "TR", "Torino": "TO", "Trapani": "TP", "Trento": "TN",
    "Treviso": "TV", "Trieste": "TS", "Udine": "UD", "Varese": "VA", "Venezia": "VE",
    "Verbano-Cusio-Ossola": "VB", "Vercelli": "VC", "Verona": "VR", "Vibo Valentia": "VV",
    "Vicenza": "VI", "Viterbo": "VT"
}

# Province aliases for flexible matching (handles variations like "Monza della Brianza" vs "Monza e Brianza")
def normalize_province_name(name: str) -> str:
    """Normalize province name to handle common variations"""
    if not name:
        return ''
    normalized = name.lower().strip()
    province_aliases = {
        'monza della brianza': 'monza e brianza',
        'monza e della brianza': 'monza e brianza',
        'monza della brienza': 'monza e brianza',  # typo variant
        'monza e brienza': 'monza e brianza',  # typo variant
        'monza-brianza': 'monza e brianza',
        'monza brianza': 'monza e brianza',
        'mb': 'monza e brianza',
        'provincia di monza e brianza': 'monza e brianza',
        'provincia di monza e della brianza': 'monza e brianza',
        'reggio nell\'emilia': 'reggio emilia',
        'reggio nell emilia': 'reggio emilia',
        'reggio-emilia': 'reggio emilia',
        're': 'reggio emilia',
        'forli-cesena': 'forlì-cesena',
        'forli cesena': 'forlì-cesena',
        'verbano cusio ossola': 'verbano-cusio-ossola',
        'vco': 'verbano-cusio-ossola',
        'pesaro urbino': 'pesaro e urbino',
        'pesaro-urbino': 'pesaro e urbino',
        'barletta andria trani': 'barletta-andria-trani',
        'bat': 'barletta-andria-trani',
        'massa carrara': 'massa-carrara',
        'massa e carrara': 'massa-carrara',
    }
    
    # Check exact match first
    if normalized in province_aliases:
        return province_aliases[normalized]
    
    # Check if contains "monza" - normalize all Monza variants
    if 'monza' in normalized:
        return 'monza e brianza'
    
    return normalized

def provincia_matches(agent_provinces: list, lead_provincia: str) -> bool:
    """Check if agent covers lead's province (with normalization)"""
    if not agent_provinces:
        return True  # Agent with no provinces covers all
    if not lead_provincia:
        return True  # Lead without province matches all agents
    
    normalized_lead = normalize_province_name(lead_provincia)
    return any(normalize_province_name(p) == normalized_lead for p in agent_provinces)

# Modelli Pydantic ed Enum: estratti in models.py (refactoring giugno 2026)
from models import *  # noqa: F401,F403

# Importazione Clienti Helper Functions
async def parse_uploaded_file(file_content: bytes, filename: str) -> ImportPreview:
    """Parse uploaded CSV/Excel file and return preview"""
    try:
        # Determine file type
        file_extension = filename.lower().split('.')[-1]
        
        if file_extension == 'csv':
            # Try different encodings for CSV
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(io.BytesIO(file_content), encoding=encoding, nrows=1000)  # Limit preview
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValueError("Unable to decode CSV file with common encodings")
                
        elif file_extension in ['xls', 'xlsx']:
            df = pd.read_excel(io.BytesIO(file_content), nrows=1000)  # Limit preview
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")
        
        # Clean column names
        df.columns = df.columns.astype(str).str.strip()
        
        # Get headers
        headers = df.columns.tolist()
        
        # Get sample data (first 5 rows)
        sample_data = []
        for _, row in df.head(5).iterrows():
            sample_data.append([str(val) if pd.notna(val) else "" for val in row.values])
        
        return ImportPreview(
            headers=headers,
            sample_data=sample_data,
            total_rows=len(df),
            file_type=file_extension
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")

def validate_cliente_data(data: dict, config: ImportConfiguration) -> tuple[bool, str]:
    """Validate cliente data according to configuration"""
    errors = []
    
    # Check required fields
    nome = data.get('nome', '').strip()
    cognome = data.get('cognome', '').strip()
    telefono = data.get('telefono', '').strip()
    
    if not nome:
        errors.append("Nome is required")
    if not cognome:
        errors.append("Cognome is required")
    if not telefono:
        errors.append("Telefono is required")
    
    # Validate phone if required
    if config.validate_phone and telefono:
        # Simple phone validation (can be enhanced)
        phone_clean = ''.join(filter(str.isdigit, telefono))
        if len(phone_clean) < 9 or len(phone_clean) > 15:
            errors.append(f"Invalid phone format: {telefono}")
    
    # Validate email if provided and validation is enabled
    email = data.get('email', '').strip()
    if config.validate_email and email:
        if '@' not in email or '.' not in email.split('@')[1]:
            errors.append(f"Invalid email format: {email}")
    
    return len(errors) == 0, "; ".join(errors)

async def process_import_batch(
    file_content: bytes,
    filename: str,
    config: ImportConfiguration,
    created_by: str
) -> ImportResult:
    """Process full import batch"""
    try:
        # Parse file
        file_extension = filename.lower().split('.')[-1]
        
        if file_extension == 'csv':
            # Try different encodings for CSV
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(io.BytesIO(file_content), encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValueError("Unable to decode CSV file")
                
        elif file_extension in ['xls', 'xlsx']:
            df = pd.read_excel(io.BytesIO(file_content))
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")
        
        # Clean column names
        df.columns = df.columns.astype(str).str.strip()
        
        # Skip header if configured
        if config.skip_header:
            df = df.iloc[1:]
        
        # Create field mapping dictionary
        field_map = {fm.csv_field: fm.client_field for fm in config.field_mappings}
        
        results = ImportResult(
            total_processed=0,
            successful=0,
            failed=0,
            errors=[],
            created_client_ids=[]
        )
        
        # Process each row
        for index, row in df.iterrows():
            try:
                results.total_processed += 1
                
                # Map fields
                cliente_data = {}
                for csv_field, client_field in field_map.items():
                    if csv_field in df.columns:
                        value = row[csv_field]
                        if pd.notna(value) and str(value).strip():
                            cliente_data[client_field] = str(value).strip()
                
                # Add required fields
                cliente_data['commessa_id'] = config.commessa_id
                cliente_data['sub_agenzia_id'] = config.sub_agenzia_id
                
                # Validate data
                is_valid, error_msg = validate_cliente_data(cliente_data, config)
                if not is_valid:
                    results.failed += 1
                    results.errors.append(f"Row {index + 1}: {error_msg}")
                    continue
                
                # Check for duplicates if configured
                if config.skip_duplicates:
                    existing = await db.clienti.find_one({
                        "telefono": cliente_data.get('telefono'),
                        "commessa_id": config.commessa_id,
                        "sub_agenzia_id": config.sub_agenzia_id
                    })
                    if existing:
                        results.failed += 1
                        results.errors.append(f"Row {index + 1}: Duplicate phone number {cliente_data.get('telefono')}")
                        continue
                
                # Create cliente
                cliente = Cliente(
                    **cliente_data,
                    created_by=created_by
                )
                
                await db.clienti.insert_one(cliente.dict())
                results.successful += 1
                results.created_client_ids.append(cliente.id)
                
            except Exception as e:
                results.failed += 1
                results.errors.append(f"Row {index + 1}: {str(e)}")
        
        return results
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import processing error: {str(e)}")

MAX_UNWORKED_LEADS_PER_AGENT = 30

async def assign_lead_to_agent(lead: Lead):
    """
    Assegna automaticamente il lead all'agente migliore basandosi su:
    1. Unit e provincia del lead
    2. Numero di lead NON GESTITI dell'agente (max 30)
    3. Performance dell'agente (chi lavora meglio riceve più lead)
    
    Un lead è considerato "non gestito" se ha ancora lo status con cui è stato assegnato.
    """
    
    # Check if the Unit has auto_assign_enabled
    if lead.unit_id:
        unit = await db.units.find_one({"id": lead.unit_id})
        if unit and not unit.get("auto_assign_enabled", True):
            # Auto-assignment disabled - assign directly to referente or agent in this Unit
            logging.info(f"[ASSIGN] Auto-assignment disabled for unit {lead.unit_id} ({unit.get('nome')}). Looking for referente or agent...")
            
            # First try to find a referente for this unit
            assignee = await db.users.find_one({
                "$or": [
                    {"unit_id": lead.unit_id},
                    {"unit_autorizzate": lead.unit_id}
                ],
                "role": "referente",
                "is_active": True
            })
            
            # If no referente found, try to find an agent
            if not assignee:
                assignee = await db.users.find_one({
                    "$or": [
                        {"unit_id": lead.unit_id},
                        {"unit_autorizzate": lead.unit_id}
                    ],
                    "role": "agente",
                    "is_active": True
                })
            
            if assignee:
                assignee_id = assignee["id"]
                assignee_name = assignee.get("username", "unknown")
                assignee_role = assignee.get("role", "unknown")
                current_esito = lead.esito or "Lead Interessato"
                
                # Assign lead
                await db.leads.update_one(
                    {"id": lead.id},
                    {
                        "$set": {
                            "assigned_agent_id": assignee_id,
                            "assigned_at": datetime.now(timezone.utc),
                            "esito_at_assignment": current_esito
                        }
                    }
                )
                
                logging.info(f"[ASSIGN] Lead {lead.id} assigned to {assignee_role} {assignee_name} ({assignee_id}) for unit {unit.get('nome')} (auto_assign disabled)")
                
                # Send email notification
                asyncio.create_task(notify_agent_new_lead(assignee_id, lead.dict()))
                
                return assignee_id
            else:
                logging.warning(f"[ASSIGN] No referente or agent found for unit {lead.unit_id} ({unit.get('nome')}). Lead {lead.id} will remain unassigned.")
                return None
    
    # Build query for agents
    query = {
        "role": "agente",
        "is_active": True
    }
    
    # IMPORTANT: Filter by unit_id if lead has one
    if lead.unit_id:
        query["unit_id"] = lead.unit_id
        logging.info(f"[ASSIGN] Looking for agents in unit_id: {lead.unit_id}")
    
    # NOTE: Province filtering is done in Python after query to support name variations
    # (e.g., "Monza della Brianza" vs "Monza e Brianza")
    logging.info(f"[ASSIGN] Looking for agents covering province: {lead.provincia}")
    
    # Find agents matching criteria
    agents = await db.users.find(query).to_list(length=None)
    
    # Filter by province using normalization (handles name variations)
    if lead.provincia:
        agents = [a for a in agents if provincia_matches(a.get("provinces", []), lead.provincia)]
        logging.info(f"[ASSIGN] After province filter: {len(agents)} agents cover province {lead.provincia}")
    
    if not agents:
        logging.warning(f"[ASSIGN] No agents found for lead {lead.id} with unit_id={lead.unit_id}, provincia={lead.provincia}")
        return None
    
    logging.info(f"[ASSIGN] Found {len(agents)} eligible agents for lead {lead.id}")
    
    # Calculate scores for each agent
    agent_scores = []
    current_esito = lead.esito or "Lead Interessato"  # Status at moment of assignment
    
    for agent in agents:
        agent_id = agent["id"]
        agent_username = agent.get("username", "unknown")
        
        # Count UNWORKED leads (leads where esito = esito_at_assignment)
        # A lead is "unworked" if its current status is the same as when it was assigned
        unworked_leads_count = await db.leads.count_documents({
            "assigned_agent_id": agent_id,
            "$or": [
                # Lead with esito_at_assignment field - compare current esito with assignment esito
                {"$expr": {"$eq": ["$esito", "$esito_at_assignment"]}},
                # Legacy: leads without esito_at_assignment - consider them unworked if esito is "Lead Interessato" or same as current
                {
                    "esito_at_assignment": {"$exists": False},
                    "esito": {"$in": [current_esito, "Lead Interessato", None, ""]}
                }
            ]
        })
        
        # Check if agent has reached max unworked leads
        if unworked_leads_count >= MAX_UNWORKED_LEADS_PER_AGENT:
            logging.info(f"[ASSIGN] Agent {agent_username} ({agent_id}) has {unworked_leads_count} unworked leads - BLOCKED (max {MAX_UNWORKED_LEADS_PER_AGENT})")
            continue  # Skip this agent
        
        # Calculate performance metrics
        # 1. Total leads worked (where status changed from assignment status)
        total_worked = await db.leads.count_documents({
            "assigned_agent_id": agent_id,
            "esito_at_assignment": {"$exists": True},
            "$expr": {"$ne": ["$esito", "$esito_at_assignment"]}
        })
        
        # 2. Average handling time (lower is better)
        agent_leads = await db.leads.find({
            "assigned_agent_id": agent_id,
            "tempo_gestione_minuti": {"$exists": True, "$ne": None, "$gt": 0}
        }).to_list(length=100)
        
        avg_handling_time = 0
        if agent_leads:
            total_time = sum([l.get("tempo_gestione_minuti", 0) for l in agent_leads])
            avg_handling_time = total_time / len(agent_leads)
        
        # 3. Conversion rate (leads closed successfully vs total assigned)
        total_assigned = await db.leads.count_documents({"assigned_agent_id": agent_id})
        successful_closures = await db.leads.count_documents({
            "assigned_agent_id": agent_id,
            "closed_at": {"$exists": True, "$ne": None}
        })
        
        conversion_rate = (successful_closures / total_assigned * 100) if total_assigned > 0 else 0
        
        # Calculate overall score (HIGHER is better for this calculation)
        # Formula: 
        # - Base score from capacity available (more room = better)
        # - Bonus for good conversion rate
        # - Bonus for fast handling time
        # - Penalty for many unworked leads
        
        capacity_available = MAX_UNWORKED_LEADS_PER_AGENT - unworked_leads_count
        capacity_score = (capacity_available / MAX_UNWORKED_LEADS_PER_AGENT) * 40  # Max 40 points
        
        conversion_score = conversion_rate * 0.3  # Max ~30 points for 100% conversion
        
        # Handling time score (faster = better, normalize to 0-20 points)
        # Assume 60 min is average, 0 min is perfect, 120+ min is bad
        if avg_handling_time == 0:
            time_score = 10  # Neutral if no data
        elif avg_handling_time <= 30:
            time_score = 20  # Excellent
        elif avg_handling_time <= 60:
            time_score = 15  # Good
        elif avg_handling_time <= 120:
            time_score = 10  # Average
        else:
            time_score = 5  # Slow
        
        total_score = capacity_score + conversion_score + time_score
        
        agent_scores.append({
            "agent_id": agent_id,
            "username": agent_username,
            "score": total_score,
            "unworked_leads": unworked_leads_count,
            "capacity_available": capacity_available,
            "total_worked": total_worked,
            "conversion_rate": conversion_rate,
            "avg_handling_time": avg_handling_time
        })
        
        logging.info(f"[ASSIGN] Agent {agent_username}: score={total_score:.1f}, unworked={unworked_leads_count}, capacity={capacity_available}, conversion={conversion_rate:.1f}%, avg_time={avg_handling_time:.0f}min")
    
    if not agent_scores:
        logging.warning(f"[ASSIGN] All agents are blocked (max leads reached). Lead {lead.id} will remain unassigned.")
        return None
    
    # Sort by score (DESCENDING - highest score wins)
    agent_scores.sort(key=lambda x: x["score"], reverse=True)
    
    selected_agent = agent_scores[0]
    selected_agent_id = selected_agent["agent_id"]
    
    logging.info(f"[ASSIGN] Selected agent: {selected_agent['username']} (score={selected_agent['score']:.1f}, unworked={selected_agent['unworked_leads']}, capacity={selected_agent['capacity_available']})")
    
    # Update lead with assignment and save the status at assignment time
    await db.leads.update_one(
        {"id": lead.id},
        {
            "$set": {
                "assigned_agent_id": selected_agent_id,
                "assigned_at": datetime.now(timezone.utc),
                "esito_at_assignment": current_esito  # Save status at assignment time
            }
        }
    )
    
    logging.info(f"[ASSIGN] Lead {lead.id} assigned to agent {selected_agent_id} ({selected_agent['username']}) with esito_at_assignment='{current_esito}'")
    
    # Send email notification to agent (async task)
    asyncio.create_task(notify_agent_new_lead(selected_agent_id, lead.dict()))
    
    return selected_agent_id


async def create_excel_report(leads_data, custom_fields_list, filename="leads_export"):
    """Create Excel file with leads data - ALL fields including custom fields"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lead Report"
    
    # Carica i nomi delle commesse e unit per la traduzione ID -> Nome
    commesse_dict = {}
    units_dict = {}
    try:
        commesse_list = await db.commesse.find({}, {"_id": 0, "id": 1, "nome": 1}).to_list(length=None)
        commesse_dict = {c["id"]: c.get("nome", c["id"]) for c in commesse_list}
        
        units_list = await db.units.find({}, {"_id": 0, "id": 1, "nome": 1}).to_list(length=None)
        units_dict = {u["id"]: u.get("nome", u["id"]) for u in units_list}
    except Exception as e:
        logging.warning(f"Error loading commesse/units names for export: {e}")
    
    # Base Headers - ALL fields from database (con nomi leggibili)
    headers = [
        "Lead ID", "Nome", "Cognome", "Telefono", "Email", "Provincia", 
        "Campagna", "Commessa", "Unit",
        "Tipologia Abitazione", "Indirizzo", "Regione", 
        "URL", "OTP", "Inserzione", "IP Address", "Contenitore",
        "Privacy Consent", "Marketing Consent", 
        "Stato/Esito", "Note", "Agente Assegnato",
        "Data Creazione", "Data Assegnazione", "Data Contatto", "Data Chiusura"
    ]
    
    # Add dynamic custom fields headers
    for custom_field in custom_fields_list:
        headers.append(f"CF: {custom_field['name']}")  # CF = Custom Field
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows - ALL fields (even if empty)
    for row, lead in enumerate(leads_data, 2):
        col = 1
        
        # Base fields - always included even if None/empty
        ws.cell(row=row, column=col, value=lead.get("lead_id", lead.get("id", "")[:8] if lead.get("id") else ""))
        col += 1
        ws.cell(row=row, column=col, value=lead.get("nome") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("cognome") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("telefono") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("email") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("provincia") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("campagna") or "")
        col += 1
        
        # Commessa: mostra nome invece di ID
        commessa_id = lead.get("commessa_id") or ""
        commessa_nome = commesse_dict.get(commessa_id, commessa_id) if commessa_id else ""
        ws.cell(row=row, column=col, value=commessa_nome)
        col += 1
        
        # Unit: mostra nome invece di ID
        unit_id = lead.get("unit_id") or lead.get("gruppo") or ""
        unit_nome = units_dict.get(unit_id, unit_id) if unit_id else ""
        ws.cell(row=row, column=col, value=unit_nome)
        col += 1
        
        ws.cell(row=row, column=col, value=lead.get("tipologia_abitazione") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("indirizzo") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("regione") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("url") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("otp") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("inserzione") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("ip_address") or "")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("contenitore") or "")
        col += 1
        
        # Consent fields - handle None separately
        privacy = lead.get("privacy_consent")
        ws.cell(row=row, column=col, value="Sì" if privacy is True else "No" if privacy is False else "Non specificato")
        col += 1
        
        marketing = lead.get("marketing_consent")
        ws.cell(row=row, column=col, value="Sì" if marketing is True else "No" if marketing is False else "Non specificato")
        col += 1
        
        ws.cell(row=row, column=col, value=lead.get("esito") or "Nuovo")
        col += 1
        ws.cell(row=row, column=col, value=lead.get("note") or "")
        col += 1
        
        # Assigned agent name (fetch from users)
        agent_id = lead.get("assigned_agent_id")
        agent_name = ""
        if agent_id:
            try:
                agent = await db.users.find_one({"id": agent_id})
                if agent:
                    agent_name = agent.get("username", "")
            except:
                pass
        ws.cell(row=row, column=col, value=agent_name)
        col += 1
        
        # Format dates - always include column even if empty
        if lead.get("created_at"):
            try:
                if isinstance(lead["created_at"], str):
                    date_obj = datetime.fromisoformat(lead["created_at"].replace("Z", "+00:00"))
                else:
                    date_obj = lead["created_at"]
                ws.cell(row=row, column=col, value=date_obj.strftime("%d/%m/%Y %H:%M"))
            except:
                ws.cell(row=row, column=col, value="")
        else:
            ws.cell(row=row, column=col, value="")
        col += 1
        
        if lead.get("assigned_at"):
            try:
                if isinstance(lead["assigned_at"], str):
                    date_obj = datetime.fromisoformat(lead["assigned_at"].replace("Z", "+00:00"))
                else:
                    date_obj = lead["assigned_at"]
                ws.cell(row=row, column=col, value=date_obj.strftime("%d/%m/%Y %H:%M"))
            except:
                ws.cell(row=row, column=col, value="")
        else:
            ws.cell(row=row, column=col, value="")
        col += 1
        
        if lead.get("contacted_at"):
            try:
                if isinstance(lead["contacted_at"], str):
                    date_obj = datetime.fromisoformat(lead["contacted_at"].replace("Z", "+00:00"))
                else:
                    date_obj = lead["contacted_at"]
                ws.cell(row=row, column=col, value=date_obj.strftime("%d/%m/%Y %H:%M"))
            except:
                ws.cell(row=row, column=col, value="")
        else:
            ws.cell(row=row, column=col, value="")
        col += 1
        
        if lead.get("closed_at"):
            try:
                if isinstance(lead["closed_at"], str):
                    date_obj = datetime.fromisoformat(lead["closed_at"].replace("Z", "+00:00"))
                else:
                    date_obj = lead["closed_at"]
                ws.cell(row=row, column=col, value=date_obj.strftime("%d/%m/%Y %H:%M"))
            except:
                ws.cell(row=row, column=col, value="")
        else:
            ws.cell(row=row, column=col, value="")
        col += 1
        
        # Custom fields - dynamically added
        custom_fields_data = lead.get("custom_fields", {})
        for custom_field in custom_fields_list:
            field_id = custom_field["id"]
            field_value = custom_fields_data.get(field_id, "")
            ws.cell(row=row, column=col, value=str(field_value) if field_value else "")
            col += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    return temp_file.name


def get_user_ip(request) -> Optional[str]:
    """Estrae l'IP dell'utente dalla richiesta"""
    # In un ambiente Kubernetes, l'IP potrebbe essere in diversi headers
    forwarded_for = getattr(request, 'headers', {}).get('X-Forwarded-For')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    
    real_ip = getattr(request, 'headers', {}).get('X-Real-IP')
    if real_ip:
        return real_ip
        
    # Fallback all'IP del client (potrebbe essere il proxy)
    client_host = getattr(request, 'client', None)
    return getattr(client_host, 'host', None) if client_host else None

async def detect_client_changes(old_client: Cliente, update_data: dict) -> List[Dict[str, str]]:
    """Rileva i cambiamenti nei dati del cliente e genera descrizioni leggibili.
    Risolve gli ID delle entità (Sub Agenzia, Commessa, Servizio, Segmento, Utente)
    nei loro nomi per una migliore leggibilità nella cronologia."""
    changes = []
    
    # Mappa dei campi con nomi user-friendly
    field_names = {
        "nome": "Nome",
        "cognome": "Cognome", 
        "email": "Email",
        "telefono": "Telefono",
        "indirizzo": "Indirizzo",
        "citta": "Città",
        "provincia": "Provincia",
        "cap": "CAP",
        "codice_fiscale": "Codice Fiscale",
        "partita_iva": "Partita IVA",
        "commessa_id": "Commessa",
        "sub_agenzia_id": "Sub Agenzia",
        "servizio_id": "Servizio",
        "tipologia_contratto": "Tipologia Contratto",
        "tipologia_contratto_id": "Tipologia Contratto",
        "segmento": "Segmento",
        "status": "Status",
        "note": "Note",
        "assigned_to": "Assegnato a",
    }

    # Helper interno per risolvere id → nome leggibile
    async def _resolve_value(field: str, value: str) -> str:
        if not value:
            return ""
        v = str(value)
        try:
            if field == "sub_agenzia_id":
                doc = await db.sub_agenzie.find_one({"id": v}, {"_id": 0, "nome": 1})
                return doc.get("nome") if doc and doc.get("nome") else v
            if field == "commessa_id":
                doc = await db.commesse.find_one({"id": v}, {"_id": 0, "nome": 1})
                return doc.get("nome") if doc and doc.get("nome") else v
            if field == "servizio_id":
                doc = await db.servizi.find_one({"id": v}, {"_id": 0, "nome": 1})
                return doc.get("nome") if doc and doc.get("nome") else v
            if field == "tipologia_contratto_id":
                doc = await db.tipologie_contratto.find_one({"id": v}, {"_id": 0, "nome": 1})
                return doc.get("nome") if doc and doc.get("nome") else v
            if field == "segmento":
                # Può essere UUID, tipo lowercase, o già nome leggibile
                if len(v) > 20:  # probabile UUID
                    doc = await db.segmenti.find_one({"id": v}, {"_id": 0, "nome": 1})
                    return doc.get("nome") if doc and doc.get("nome") else v
                # Se è una stringa breve, usala direttamente con prima lettera maiuscola
                return v[0].upper() + v[1:] if v else v
            if field == "assigned_to":
                doc = await db.users.find_one({"id": v}, {"_id": 0, "username": 1})
                return doc.get("username") if doc and doc.get("username") else v
        except Exception:
            return v
        return v

    for field, new_value in update_data.items():
        if field in ["updated_at", "dati_aggiuntivi"]:  # Skip meta fields
            continue
            
        old_value = getattr(old_client, field, None)
        
        # Convert values to string for comparison
        old_str = str(old_value) if old_value is not None else ""
        new_str = str(new_value) if new_value is not None else ""
        
        if old_str != new_str:
            field_display = field_names.get(field, field.title())
            # Risolvi i valori in nomi leggibili (per i campi mappati)
            old_display = await _resolve_value(field, old_str) if old_str else ""
            new_display = await _resolve_value(field, new_str) if new_str else ""
            changes.append({
                "field": field,
                "field_display": field_display,
                "old_value": old_display,
                "new_value": new_display,
                "description": f"{field_display} modificato da '{old_display}' a '{new_display}'"
            })
    
    return changes


async def _expand_segmento_filter_values(values: list[str]) -> list[str]:
    """Given a list of segmento filter values (tipo like 'privato'/'business' or UUIDs),
    return the expanded list that ALSO includes all UUIDs of segmenti having that tipo
    AND the case variants of tipo strings (privato/Privato/PRIVATO).

    Fix for legacy data: clienti.segmento is stored sometimes as the tipo string
    in different casings ('privato'/'Privato'/'PRIVATO') and sometimes as the segmento UUID.
    """
    if not values:
        return values
    expanded: set[str] = set()
    tipo_values: set[str] = set()
    for v in values:
        if v is None:
            continue
        v = str(v).strip()
        if not v:
            continue
        expanded.add(v)
        # If value is a known "tipo" string, also add all UUIDs of segmenti with that tipo
        # AND the case variants (lowercase, capitalized, uppercase)
        if v.lower() in ("privato", "business"):
            tipo_lower = v.lower()
            tipo_values.add(tipo_lower)
            expanded.add(tipo_lower)
            expanded.add(tipo_lower.capitalize())
            expanded.add(tipo_lower.upper())
    if tipo_values:
        segmenti_cursor = db.segmenti.find(
            {"tipo": {"$in": list(tipo_values)}}, {"_id": 0, "id": 1, "nome": 1}
        )
        async for s in segmenti_cursor:
            sid = s.get("id")
            if sid:
                expanded.add(sid)
            # Anche il nome del segmento, perché ora il frontend salva il nome
            sname = s.get("nome")
            if sname:
                expanded.add(sname)
    return list(expanded)



# Gestione Clienti

async def create_clienti_excel_report(clienti_data, filename="clienti_export", custom_fields=None):
    """Create Excel file with clienti data - ALL fields included (standard + custom), one row per SIM.
    
    Args:
        clienti_data: list of cliente dicts (may contain 'dati_aggiuntivi' key)
        filename: output filename
        custom_fields: optional list of dicts with 'name' and 'label' keys. If None, the function
                       auto-fetches all active custom fields from the `cliente_custom_fields` collection.
    """
    # Auto-fetch custom fields if not provided
    if custom_fields is None:
        try:
            raw = await db.cliente_custom_fields.find({"active": True}, {"_id": 0}).to_list(length=None)
            # De-duplicate by name (same logical field may exist across commessa/tipologia combos — use the first label seen)
            seen = {}
            for f in raw:
                nm = f.get("name")
                if nm and nm not in seen:
                    seen[nm] = f.get("label") or nm
            custom_fields = [{"name": k, "label": v} for k, v in sorted(seen.items(), key=lambda x: x[1].lower())]
        except Exception as e:
            logging.warning(f"Could not fetch custom_fields for export: {e}")
            custom_fields = []
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clienti Report"
    
    # Headers - ALL FIELDS from cliente model
    headers = [
        # Dati Identificativi
        "ID Cliente", "Numero Ordine", "Account", "Cliente ID",
        # Dati Anagrafici
        "Nome", "Cognome", "Ragione Sociale", "Data Nascita", "Luogo Nascita",
        "Comune Residenza", "Provincia Residenza", "Genere",
        # Contatti
        "Email", "Telefono", "Cellulare",
        # Indirizzo
        "Indirizzo", "Numero Civico", "Comune", "Provincia", "CAP",
        # Dati Fiscali
        "Codice Fiscale", "Partita IVA",
        # Documento
        "Tipo Documento", "Numero Documento", "Data Rilascio", "Luogo Rilascio", "Scadenza Documento",
        # Dati Organizzativi
        "Sub Agenzia", "Commessa", "Servizio", "Tipologia Contratto", "Segmento", "Offerta",
        # Telefonia Fastweb
        "Tecnologia", "Codice Migrazione", "Gestore", "Convergenza",
        # Energia Fastweb
        "Tipologia Energia", "Codice POD", "Consumo Annuo", "Potenza Contatore", "Potenza Impegnata", "Fornitore Attuale",
        # Telepass
        "OBU",
        # Modalità Pagamento
        "Modalità Pagamento", "IBAN", "Intestatario Diverso (IBAN)",
        "Numero Carta", "Intestatario Carta", "CVV", "Mese Scadenza", "Anno Scadenza",
        # SIM Info (Convergenza or Mobile)
        "Tipo SIM", "Numero SIM", "Numero Cellulare SIM", "ICCID SIM", "Operatore SIM",
        "Telefono da Portare", "Titolare Diverso", "Offerta SIM", "Utente Assegnato SIM",
        # System Fields
        "Status", "Utente Creatore", "Data Creazione", "Note", "Note Back Office",
        # Ultime note dallo storico (immutabile) per ogni tipologia
        "Ultima Nota Cliente", "Ultima Nota Cliente - Autore", "Ultima Nota Cliente - Data",
        "Ultima Nota Back Office", "Ultima Nota Back Office - Autore", "Ultima Nota Back Office - Data",
        "Ultima Nota Post Vendita", "Ultima Nota Post Vendita - Autore", "Ultima Nota Post Vendita - Data",
        # Post Vendita
        "Post Vendita - In Workflow", "Post Vendita - Stato", "Post Vendita - Esito (Stage)", "Post Vendita - Ultimo Aggiornamento", "Codice Account"
    ]
    
    # Append custom field columns (dynamic, from cliente_custom_fields)
    for cf in (custom_fields or []):
        headers.append(f"[Custom] {cf['label']}")
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Pre-load the LATEST note per (cliente_id, tipo) from the immutable history
    # so each row in the export shows the most recent cliente / backoffice / post_vendita note
    cliente_ids_for_notes = [c.get("id") for c in clienti_data if c.get("id")]
    latest_notes_map: Dict[str, Dict[str, dict]] = {}
    if cliente_ids_for_notes:
        notes_pipeline = [
            {"$match": {"cliente_id": {"$in": cliente_ids_for_notes}, "tipo": {"$in": ["cliente", "backoffice", "post_vendita"]}}},
            {"$sort": {"created_at": -1}},
            {"$group": {
                "_id": {"cliente_id": "$cliente_id", "tipo": "$tipo"},
                "content": {"$first": "$content"},
                "created_at": {"$first": "$created_at"},
                "created_by_username": {"$first": "$created_by_username"},
            }},
        ]
        async for entry in db.cliente_note_history.aggregate(notes_pipeline):
            cid = entry["_id"]["cliente_id"]
            tipo = entry["_id"]["tipo"]
            latest_notes_map.setdefault(cid, {})[tipo] = {
                "content": entry.get("content", ""),
                "created_at": entry.get("created_at"),
                "created_by_username": entry.get("created_by_username", ""),
            }

    def _fmt_dt(v):
        if not v:
            return ""
        if isinstance(v, str):
            return v
        try:
            return v.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return str(v)
    
    # Data rows - ALL FIELDS
    for row_idx, cliente in enumerate(clienti_data, 2):
        col = 1
        
        # Dati Identificativi
        ws.cell(row=row_idx, column=col, value=cliente.get("id", "")[:8]); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("numero_ordine", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("account", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("cliente_id", "")); col += 1
        
        # Dati Anagrafici
        ws.cell(row=row_idx, column=col, value=cliente.get("nome", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("cognome", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("ragione_sociale", "")); col += 1
        
        # Format date of birth
        data_nascita = cliente.get("data_nascita")
        if data_nascita:
            if isinstance(data_nascita, str):
                ws.cell(row=row_idx, column=col, value=data_nascita)
            else:
                ws.cell(row=row_idx, column=col, value=data_nascita.strftime("%d/%m/%Y") if data_nascita else "")
        else:
            ws.cell(row=row_idx, column=col, value="")
        col += 1
        
        ws.cell(row=row_idx, column=col, value=cliente.get("luogo_nascita", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("comune_residenza", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("provincia_residenza", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("genere", "")); col += 1
        
        # Contatti
        ws.cell(row=row_idx, column=col, value=cliente.get("email", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("telefono", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("cellulare", "") or cliente.get("telefono2", "")); col += 1
        
        # Indirizzo
        ws.cell(row=row_idx, column=col, value=cliente.get("indirizzo", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("numero_civico", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("comune", "") or cliente.get("citta", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("provincia", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("cap", "")); col += 1
        
        # Dati Fiscali
        ws.cell(row=row_idx, column=col, value=cliente.get("codice_fiscale", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("partita_iva", "")); col += 1
        
        # Documento
        ws.cell(row=row_idx, column=col, value=cliente.get("tipo_documento", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("numero_documento", "")); col += 1
        
        # Format document date
        data_rilascio = cliente.get("data_rilascio")
        if data_rilascio:
            if isinstance(data_rilascio, str):
                ws.cell(row=row_idx, column=col, value=data_rilascio)
            else:
                ws.cell(row=row_idx, column=col, value=data_rilascio.strftime("%d/%m/%Y") if data_rilascio else "")
        else:
            ws.cell(row=row_idx, column=col, value="")
        col += 1
        
        ws.cell(row=row_idx, column=col, value=cliente.get("luogo_rilascio", "")); col += 1
        
        # Format document expiry date
        scadenza_documento = cliente.get("scadenza_documento")
        if scadenza_documento:
            if isinstance(scadenza_documento, str):
                ws.cell(row=row_idx, column=col, value=scadenza_documento)
            else:
                ws.cell(row=row_idx, column=col, value=scadenza_documento.strftime("%d/%m/%Y") if scadenza_documento else "")
        else:
            ws.cell(row=row_idx, column=col, value="")
        col += 1
        
        # Dati Organizzativi
        ws.cell(row=row_idx, column=col, value=cliente.get("sub_agenzia_name", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("commessa_name", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("servizio_name", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("tipologia_contratto_display", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("segmento_display", "")); col += 1
        
        # Offerta: Use SIM offerta for SIM rows, cliente offerta for fixed line rows
        sim_type = cliente.get("sim_type", "")
        if sim_type in ["SIM Convergenza", "Mobile"]:
            # For SIM rows, show the SIM-specific offerta
            ws.cell(row=row_idx, column=col, value=cliente.get("sim_offerta_name", "")); col += 1
        else:
            # For fixed line rows (or no convergenza), show the cliente offerta
            ws.cell(row=row_idx, column=col, value=cliente.get("offerta_name", "")); col += 1
        
        # Telefonia Fastweb
        ws.cell(row=row_idx, column=col, value=cliente.get("tecnologia", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("codice_migrazione", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("gestore", "")); col += 1
        ws.cell(row=row_idx, column=col, value="Sì" if cliente.get("convergenza") else "No"); col += 1
        
        # Energia Fastweb
        ws.cell(row=row_idx, column=col, value=cliente.get("energia_tipologia", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("codice_pod", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("energia_consumo_annuo", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("energia_potenza_contatore", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("energia_potenza_impegnata", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("energia_fornitore_attuale", "")); col += 1
        
        # Telepass
        ws.cell(row=row_idx, column=col, value=cliente.get("obu", "")); col += 1
        
        # Modalità Pagamento
        modalita_pagamento_display = {
            'iban': 'IBAN',
            'carta_credito': 'Carta di Credito'
        }.get(cliente.get("modalita_pagamento", ""), cliente.get("modalita_pagamento", ""))
        ws.cell(row=row_idx, column=col, value=modalita_pagamento_display); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("iban", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("intestatario_diverso", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("numero_carta", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("intestatario_carta", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("cvv_carta", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("mese_carta", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("anno_carta", "")); col += 1
        
        # SIM Info
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_type", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_index", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_numero_cellulare", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_iccid", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_operatore", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_telefono_da_portare", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_titolare_diverso", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_offerta_name", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("sim_assigned_user", "")); col += 1  # NEW: Assigned user
        
        # System Fields
        ws.cell(row=row_idx, column=col, value=cliente.get("status", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("created_by_name", "")); col += 1
        
        # Format creation date
        created_at = cliente.get("created_at")
        if created_at:
            if isinstance(created_at, str):
                ws.cell(row=row_idx, column=col, value=created_at)
            else:
                ws.cell(row=row_idx, column=col, value=created_at.strftime("%d/%m/%Y %H:%M") if created_at else "")
        else:
            ws.cell(row=row_idx, column=col, value="")
        col += 1
        
        ws.cell(row=row_idx, column=col, value=cliente.get("note", "")); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("note_backoffice", "") or cliente.get("note_back_office", "")); col += 1
        
        # Ultime note dallo storico immutabile
        _cnotes = latest_notes_map.get(cliente.get("id"), {})
        for _tipo in ("cliente", "backoffice", "post_vendita"):
            _entry = _cnotes.get(_tipo) or {}
            ws.cell(row=row_idx, column=col, value=_entry.get("content", "")); col += 1
            ws.cell(row=row_idx, column=col, value=_entry.get("created_by_username", "")); col += 1
            ws.cell(row=row_idx, column=col, value=_fmt_dt(_entry.get("created_at"))); col += 1
        
        # Post Vendita
        ws.cell(row=row_idx, column=col, value="Sì" if cliente.get("passed_to_post_vendita") else "No"); col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("post_vendita_status_label") or cliente.get("post_vendita_status") or ""); col += 1
        _pv_stage = cliente.get("post_vendita_stage") or ""
        _pv_stage_label = {"attivato": "🟢 Attivato", "ko": "🔴 KO", "lavorazione": "🟡 In Lavorazione"}.get(_pv_stage, "")
        ws.cell(row=row_idx, column=col, value=_pv_stage_label); col += 1
        _pv_ts = cliente.get("post_vendita_status_updated_at")
        if _pv_ts:
            if isinstance(_pv_ts, str):
                ws.cell(row=row_idx, column=col, value=_pv_ts)
            else:
                ws.cell(row=row_idx, column=col, value=_pv_ts.strftime("%d/%m/%Y %H:%M"))
        else:
            ws.cell(row=row_idx, column=col, value="")
        col += 1
        ws.cell(row=row_idx, column=col, value=cliente.get("codice_account", "") or ""); col += 1
        
        # Custom fields values (from dati_aggiuntivi)
        dati_agg = cliente.get("dati_aggiuntivi") or {}
        for cf in (custom_fields or []):
            v = dati_agg.get(cf["name"], "")
            # Format lists (multi_select) as comma-separated
            if isinstance(v, list):
                v = ", ".join(str(x) for x in v)
            elif isinstance(v, bool):
                v = "Sì" if v else "No"
            ws.cell(row=row_idx, column=col, value=v if v is not None else ""); col += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Max width 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    return temp_file.name



# Elementi hardcoded commesse (spostati da server.py - fase 3)
async def get_hardcoded_tipologie_contratto():
    """Helper function to get hardcoded tipologie contratto"""
    # Lista base per tutti i servizi di Fastweb
    tipologie_base = [
        {"value": "energia_fastweb", "label": "Energia Fastweb"},
        {"value": "telefonia_fastweb", "label": "Telefonia Fastweb"}
    ]
    
    # Tipologie aggiuntive per servizi specifici
    tipologie_aggiuntive = [
        {"value": "ho_mobile", "label": "Ho Mobile"},
        {"value": "telepass", "label": "Telepass"}
    ]
    
    return tipologie_base + tipologie_aggiuntive


async def should_use_hardcoded_elements():
    """Helper function to check if hardcoded elements should be used"""
    try:
        setting = await db.system_settings.find_one({"key": "hardcoded_elements_disabled"})
        return not (setting and setting.get("value", False))
    except:
        return True  # Default to using hardcoded if check fails
