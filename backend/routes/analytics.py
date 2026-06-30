"""Route: Analytics agenti/supervisor/referenti, export Excel lead, pivot — estratte da server.py (refactoring fase 3, giugno 2026)."""
import asyncio
import io
import json
import logging
import os
import re
import uuid
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Dict, Any

from fastapi import (
    APIRouter, HTTPException, Depends, Query, Body, Request,
    UploadFile, File, Form, status,
)
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse, Response

from database import db
from security import (
    get_current_user, get_password_hash, verify_password, create_access_token,
    pwd_context, ACCESS_TOKEN_EXPIRE_MINUTES,
    get_user_commessa_authorizations, check_commessa_access, get_user_accessible_commesse,
    get_user_accessible_sub_agenzie, can_user_access_cliente, can_user_access_cliente_notes,
    can_user_delete_cliente, can_user_modify_cliente, can_user_access_document,
    get_user_accessible_documents,
)
from helpers import (
    ITALIAN_PROVINCES, PROVINCE_TO_CODE, normalize_province_name, provincia_matches,
    MAX_UNWORKED_LEADS_PER_AGENT, assign_lead_to_agent,
    parse_uploaded_file, validate_cliente_data, process_import_batch,
    create_excel_report, create_clienti_excel_report,
    get_user_ip, detect_client_changes, _expand_segmento_filter_values,
)
from services import (
    UPLOAD_DIR, MAX_FILE_SIZE, ALLOWED_FILE_TYPES,
    aruba_service, validate_uploaded_file, save_temporary_file, create_document_record,
)
from notifications import notify_agent_new_lead, send_email_notification
from audit import log_client_action
from models import *  # noqa: F401,F403
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()
logger = logging.getLogger(__name__)

@router.get("/analytics/agent/{agent_id}")
async def get_agent_analytics(
    agent_id: str, 
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Permission check
    if current_user.role == UserRole.AGENTE and current_user.id != agent_id:
        raise HTTPException(status_code=403, detail="Can only view your own analytics")
    elif current_user.role == UserRole.REFERENTE:
        # Check if agent is under this referente
        agent = await db.users.find_one({"id": agent_id})
        if not agent or agent["referente_id"] != current_user.id:
            raise HTTPException(status_code=403, detail="Can only view analytics for your agents")
    elif current_user.role == UserRole.SUPERVISOR:
        # Supervisor can view analytics for agents in their authorized Units
        agent = await db.users.find_one({"id": agent_id})
        supervisor_units = (current_user.unit_autorizzate or []) + ([current_user.unit_id] if current_user.unit_id else [])
        if not agent or agent.get("unit_id") not in supervisor_units:
            raise HTTPException(status_code=403, detail="Puoi vedere analytics solo degli agenti nelle tue Unit")
    elif current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    # Get agent info
    agent = await db.users.find_one({"id": agent_id})
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    
    # Build base query with date filters
    base_query = {"assigned_agent_id": agent_id}
    
    # Add date filters if provided
    if date_from or date_to:
        date_filter = {}
        if date_from:
            try:
                date_from_obj, _ = __import__("helpers", fromlist=["rome_date_to_utc_range"]).rome_date_to_utc_range(date_from, current_user.timezone)
                date_filter["$gte"] = date_from_obj
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date_from format. Use YYYY-MM-DD")
        if date_to:
            try:
                _, date_to_obj = __import__("helpers", fromlist=["rome_date_to_utc_range"]).rome_date_to_utc_range(date_to, current_user.timezone)
                date_filter["$lte"] = date_to_obj
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date_to format. Use YYYY-MM-DD")
        base_query["created_at"] = date_filter
    
    # Get agent's leads statistics
    total_leads = await db.leads.count_documents(base_query)
    
    # Contacted leads = leads with esito that is NOT "Nuovo" (null, empty, or "Nuovo" string)
    # Only leads that have been worked on (changed from Nuovo to another status) count as contacted
    contacted_query = {**base_query, "$and": [
        {"esito": {"$exists": True}},
        {"esito": {"$ne": None}},
        {"esito": {"$ne": ""}},
        {"esito": {"$ne": "Nuovo"}}
    ]}
    contacted_leads = await db.leads.count_documents(contacted_query)
    
    # Leads by outcome - COUNT ALL ACTUAL VALUES IN DATABASE
    outcomes = {}
    
    # Use MongoDB aggregation to get ALL distinct esito values with counts
    pipeline = [
        {"$match": base_query},
        {"$group": {
            "_id": "$esito",
            "count": {"$sum": 1}
        }},
        {"$sort": {"count": -1}}
    ]
    
    esito_counts = await db.leads.aggregate(pipeline).to_list(length=None)
    
    for item in esito_counts:
        esito_value = item["_id"]
        count = item["count"]
        
        # Handle None/empty/Nuovo esito as "Nuovo"
        # This includes: null, empty string, "Nuovo" string, and non-existent field
        if esito_value is None or esito_value == "" or not esito_value or esito_value == "Nuovo":
            outcomes["Nuovo"] = outcomes.get("Nuovo", 0) + count
        else:
            outcomes[esito_value] = count
    
    # Leads this week/month - respect date filters if provided
    now = datetime.now(timezone.utc)
    week_start = now.replace(hour=0, minute=0, second=0) - timedelta(days=7)
    month_start = now.replace(day=1, hour=0, minute=0, second=0)
    
    week_query = {**base_query}
    # If no date filters provided, use week_start; otherwise use existing filter
    if "created_at" not in base_query:
        week_query["created_at"] = {"$gte": week_start}
    
    month_query = {**base_query}
    # If no date filters provided, use month_start; otherwise use existing filter
    if "created_at" not in base_query:
        month_query["created_at"] = {"$gte": month_start}
    
    leads_this_week = await db.leads.count_documents(week_query)
    leads_this_month = await db.leads.count_documents(month_query)
    
    return {
        "agent": {
            "id": agent["id"],
            "username": agent["username"],
            "email": agent["email"]
        },
        "stats": {
            "total_leads": total_leads,
            "contacted_leads": contacted_leads,
            "contact_rate": round((contacted_leads / total_leads * 100) if total_leads > 0 else 0, 2),
            "leads_this_week": leads_this_week,
            "leads_this_month": leads_this_month,
            "outcomes": outcomes
        }
    }

@router.get("/analytics/supervisor/unit")
async def get_supervisor_unit_analytics(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get analytics for the Supervisor's Units - includes all agents and referenti"""
    # Permission check - only Supervisor and Admin
    if current_user.role == UserRole.SUPERVISOR:
        supervisor_units = current_user.unit_autorizzate or []
        if current_user.unit_id and current_user.unit_id not in supervisor_units:
            supervisor_units.append(current_user.unit_id)
        if not supervisor_units:
            raise HTTPException(status_code=403, detail="Supervisor non ha Unit assegnate")
    elif current_user.role == UserRole.ADMIN:
        raise HTTPException(status_code=400, detail="Admin deve usare altri endpoint con unit_id specifico")
    else:
        raise HTTPException(status_code=403, detail="Solo Supervisor può accedere a questo endpoint")
    
    # Get all Units info
    units_info = await db.units.find({"id": {"$in": supervisor_units}}).to_list(length=None)
    unit_names = {u["id"]: u.get("nome", u["id"]) for u in units_info}
    
    # Get all agents and referenti in these Units
    users_in_units = await db.users.find({
        "unit_id": {"$in": supervisor_units},
        "role": {"$in": ["agente", "referente"]},
        "is_active": True
    }).to_list(length=None)
    
    agents = [u for u in users_in_units if u.get("role") == "agente"]
    referenti = [u for u in users_in_units if u.get("role") == "referente"]
    
    # Build date filters
    date_filter = {}
    if date_from:
        try:
            date_from_obj, _ = __import__("helpers", fromlist=["rome_date_to_utc_range"]).rome_date_to_utc_range(date_from, current_user.timezone)
            date_filter["$gte"] = date_from_obj
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format. Use YYYY-MM-DD")
    if date_to:
        try:
            _, date_to_obj = __import__("helpers", fromlist=["rome_date_to_utc_range"]).rome_date_to_utc_range(date_to, current_user.timezone)
            date_filter["$lte"] = date_to_obj
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format. Use YYYY-MM-DD")
    
    # Base query for leads in supervisor's units
    base_query = {"unit_id": {"$in": supervisor_units}}
    if date_filter:
        base_query["created_at"] = date_filter
    
    # Unit-level stats
    total_leads = await db.leads.count_documents(base_query)
    
    # Contacted leads
    contacted_query = {**base_query, "$and": [
        {"esito": {"$exists": True}},
        {"esito": {"$ne": None}},
        {"esito": {"$ne": ""}},
        {"esito": {"$ne": "Nuovo"}}
    ]}
    contacted_leads = await db.leads.count_documents(contacted_query)
    
    # Unassigned leads
    unassigned_query = {**base_query, "$or": [
        {"assigned_agent_id": None},
        {"assigned_agent_id": {"$exists": False}}
    ]}
    unassigned_leads = await db.leads.count_documents(unassigned_query)
    
    # Per-agent stats
    agent_stats = []
    all_outcomes = set()  # Collect all unique outcomes
    
    for agent in agents:
        agent_query = {"assigned_agent_id": agent["id"]}
        if date_filter:
            agent_query["created_at"] = date_filter
        
        agent_total = await db.leads.count_documents(agent_query)
        agent_contacted_query = {**agent_query, "$and": [
            {"esito": {"$exists": True}},
            {"esito": {"$ne": None}},
            {"esito": {"$ne": ""}},
            {"esito": {"$ne": "Nuovo"}}
        ]}
        agent_contacted = await db.leads.count_documents(agent_contacted_query)
        
        # Get outcomes breakdown for this agent
        agent_outcomes = {}
        outcome_pipeline = [
            {"$match": agent_query},
            {"$group": {"_id": "$esito", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        outcome_results = await db.leads.aggregate(outcome_pipeline).to_list(length=None)
        for item in outcome_results:
            esito = item["_id"] if item["_id"] else "Nuovo"
            if esito == "" or esito is None:
                esito = "Nuovo"
            agent_outcomes[esito] = item["count"]
            all_outcomes.add(esito)
        
        agent_stats.append({
            "id": agent["id"],
            "username": agent.get("username"),
            "email": agent.get("email"),
            "referente_id": agent.get("referente_id"),
            "unit_id": agent.get("unit_id"),
            "unit_nome": unit_names.get(agent.get("unit_id"), "N/A"),
            "total_leads": agent_total,
            "contacted_leads": agent_contacted,
            "contact_rate": round((agent_contacted / agent_total * 100) if agent_total > 0 else 0, 2),
            "outcomes": agent_outcomes
        })
    
    # Calculate total outcomes across all agents
    total_outcomes = {}
    for agent in agent_stats:
        for esito, count in agent.get("outcomes", {}).items():
            total_outcomes[esito] = total_outcomes.get(esito, 0) + count
    
    # Per-referente stats
    referente_stats = []
    for referente in referenti:
        # Get agents under this referente
        ref_agents = [a for a in agents if a.get("referente_id") == referente["id"]]
        ref_agent_ids = [a["id"] for a in ref_agents]
        ref_agent_ids.append(referente["id"])  # Include referente's own leads
        
        ref_query = {"assigned_agent_id": {"$in": ref_agent_ids}}
        if date_filter:
            ref_query["created_at"] = date_filter
        
        ref_total = await db.leads.count_documents(ref_query)
        ref_contacted_query = {**ref_query, "$and": [
            {"esito": {"$exists": True}},
            {"esito": {"$ne": None}},
            {"esito": {"$ne": ""}},
            {"esito": {"$ne": "Nuovo"}}
        ]}
        ref_contacted = await db.leads.count_documents(ref_contacted_query)
        
        referente_stats.append({
            "id": referente["id"],
            "username": referente.get("username"),
            "email": referente.get("email"),
            "unit_id": referente.get("unit_id"),
            "unit_nome": unit_names.get(referente.get("unit_id"), "N/A"),
            "agents_count": len(ref_agents),
            "total_leads": ref_total,
            "contacted_leads": ref_contacted,
            "contact_rate": round((ref_contacted / ref_total * 100) if ref_total > 0 else 0, 2)
        })
    
    return {
        "units": [{"id": u_id, "nome": u_name} for u_id, u_name in unit_names.items()],
        "stats": {
            "total_leads": total_leads,
            "contacted_leads": contacted_leads,
            "unassigned_leads": unassigned_leads,
            "contact_rate": round((contacted_leads / total_leads * 100) if total_leads > 0 else 0, 2),
            "total_agents": len(agents),
            "total_referenti": len(referenti),
            "total_units": len(supervisor_units)
        },
        "agents": agent_stats,
        "referenti": referente_stats,
        "outcomes": total_outcomes
    }

@router.get("/analytics/referente/{referente_id}")
async def get_referente_analytics(
    referente_id: str, 
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Permission check
    if current_user.role == UserRole.ADMIN:
        pass  # Admin can access any referente analytics
    elif current_user.role == UserRole.REFERENTE:
        # Referente can only view their own analytics
        if current_user.id != referente_id:
            raise HTTPException(status_code=403, detail="Can only view your own analytics")
    elif current_user.role == UserRole.SUPER_REFERENTE:
        # Super Referente can view analytics for their authorized referenti
        if referente_id not in (current_user.referenti_autorizzati or []):
            raise HTTPException(status_code=403, detail="Puoi vedere analytics solo dei referenti autorizzati")
    elif current_user.role == UserRole.SUPERVISOR:
        # Supervisor can view analytics for referenti in their authorized Units
        referente = await db.users.find_one({"id": referente_id})
        supervisor_units = (current_user.unit_autorizzate or []) + ([current_user.unit_id] if current_user.unit_id else [])
        if not referente or referente.get("unit_id") not in supervisor_units:
            raise HTTPException(status_code=403, detail="Puoi vedere analytics solo dei referenti nelle tue Unit")
    else:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    # Get referente info
    referente = await db.users.find_one({"id": referente_id})
    if not referente:
        raise HTTPException(status_code=404, detail="Referente not found")
    
    # Get all agents under this referente
    agents = await db.users.find({"referente_id": referente_id}).to_list(length=None)
    agent_ids = [agent["id"] for agent in agents]
    
    # Build base query with date filters
    base_query = {"assigned_agent_id": {"$in": agent_ids}}
    
    # For Super Referente: also filter by their Unit
    if current_user.role == UserRole.SUPER_REFERENTE and current_user.unit_id:
        base_query["unit_id"] = current_user.unit_id
        logging.info(f"[ANALYTICS] Super Referente {current_user.username} filtering analytics for unit {current_user.unit_id}")
    
    # Add date filters if provided
    if date_from or date_to:
        date_filter = {}
        if date_from:
            try:
                date_from_obj, _ = __import__("helpers", fromlist=["rome_date_to_utc_range"]).rome_date_to_utc_range(date_from, current_user.timezone)
                date_filter["$gte"] = date_from_obj
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date_from format. Use YYYY-MM-DD")
        if date_to:
            try:
                _, date_to_obj = __import__("helpers", fromlist=["rome_date_to_utc_range"]).rome_date_to_utc_range(date_to, current_user.timezone)
                date_filter["$lte"] = date_to_obj
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date_to format. Use YYYY-MM-DD")
        base_query["created_at"] = date_filter
    
    # Aggregate statistics for all agents under referente
    total_leads = await db.leads.count_documents(base_query)
    
    # Contacted leads = leads with esito that is NOT "Nuovo" (null, empty, or "Nuovo" string)
    contacted_query = {**base_query, "$and": [
        {"esito": {"$exists": True}},
        {"esito": {"$ne": None}},
        {"esito": {"$ne": ""}},
        {"esito": {"$ne": "Nuovo"}}
    ]}
    contacted_leads = await db.leads.count_documents(contacted_query)
    
    # Per-agent breakdown - respect date filters AND unit filter for Super Referente
    agent_stats = []
    for agent in agents:
        agent_base_query = {"assigned_agent_id": agent["id"]}
        
        # Add unit filter for Super Referente
        if "unit_id" in base_query:
            agent_base_query["unit_id"] = base_query["unit_id"]
        
        # Add same date filters as parent query
        if "created_at" in base_query:
            agent_base_query["created_at"] = base_query["created_at"]
        
        agent_leads = await db.leads.count_documents(agent_base_query)
        
        # Contacted leads = leads with esito that is NOT "Nuovo"
        agent_contacted_query = {**agent_base_query, "$and": [
            {"esito": {"$exists": True}},
            {"esito": {"$ne": None}},
            {"esito": {"$ne": ""}},
            {"esito": {"$ne": "Nuovo"}}
        ]}
        agent_contacted = await db.leads.count_documents(agent_contacted_query)
        
        # Get outcomes for this agent
        agent_outcomes = {}
        agent_pipeline = [
            {"$match": agent_base_query},
            {"$group": {
                "_id": "$esito",
                "count": {"$sum": 1}
            }},
            {"$sort": {"count": -1}}
        ]
        
        agent_esito_counts = await db.leads.aggregate(agent_pipeline).to_list(length=None)
        
        for item in agent_esito_counts:
            esito_value = item["_id"]
            count = item["count"]
            
            # Handle None/empty/Nuovo esito as "Nuovo"
            if esito_value is None or esito_value == "" or not esito_value or esito_value == "Nuovo":
                agent_outcomes["Nuovo"] = agent_outcomes.get("Nuovo", 0) + count
            else:
                agent_outcomes[esito_value] = count
        
        agent_stats.append({
            "agent": {
                "id": agent["id"],
                "username": agent["username"],
                "email": agent["email"]
            },
            "total_leads": agent_leads,
            "contacted_leads": agent_contacted,
            "contact_rate": round((agent_contacted / agent_leads * 100) if agent_leads > 0 else 0, 2),
            "outcomes": agent_outcomes
        })
    
    # Leads by outcome - COUNT ALL ACTUAL VALUES IN DATABASE (same as agent analytics)
    outcomes = {}
    
    # Use MongoDB aggregation to get ALL distinct esito values with counts for all agents
    pipeline = [
        {"$match": base_query},
        {"$group": {
            "_id": "$esito",
            "count": {"$sum": 1}
        }},
        {"$sort": {"count": -1}}
    ]
    
    esito_counts = await db.leads.aggregate(pipeline).to_list(length=None)
    
    for item in esito_counts:
        esito_value = item["_id"]
        count = item["count"]
        
        # Handle None/empty/Nuovo esito as "Nuovo"
        if esito_value is None or esito_value == "" or not esito_value or esito_value == "Nuovo":
            outcomes["Nuovo"] = outcomes.get("Nuovo", 0) + count
        else:
            outcomes[esito_value] = count
    
    return {
        "referente": {
            "id": referente["id"],
            "username": referente["username"],
            "email": referente["email"]
        },
        "total_agents": len(agents),
        "total_stats": {
            "total_leads": total_leads,
            "contacted_leads": contacted_leads,
            "contact_rate": round((contacted_leads / total_leads * 100) if total_leads > 0 else 0, 2)
        },
        "outcomes": outcomes,  # NEW: Distribution of lead outcomes
        "agent_breakdown": agent_stats
    }

# Excel Export System
@router.get("/leads/export")
async def export_leads_excel(
    unit_id: Optional[str] = None,
    campagna: Optional[str] = None,
    provincia: Optional[str] = None,
    status: Optional[str] = None,  # NEW: Status filter
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    assigned_agent_id: Optional[str] = None,  # NEW: Agent filter
    search: Optional[str] = None,  # NEW: Search filter
    current_user: User = Depends(get_current_user)
):
    """Export leads to Excel file with ALL filters and ALL fields including custom fields"""
    query = {}
    
    # Role-based filtering (same as get_leads)
    if current_user.role == UserRole.AGENTE:
        query["assigned_agent_id"] = current_user.id
    elif current_user.role == UserRole.REFERENTE:
        agents = await db.users.find({"referente_id": current_user.id}).to_list(length=None)
        agent_ids = [agent["id"] for agent in agents]
        agent_ids.append(current_user.id)
        query["assigned_agent_id"] = {"$in": agent_ids}
    elif current_user.role == UserRole.SUPERVISOR:
        # Supervisor can export ALL leads from their authorized Units
        supervisor_units = current_user.unit_autorizzate or []
        if current_user.unit_id and current_user.unit_id not in supervisor_units:
            supervisor_units.append(current_user.unit_id)
        if supervisor_units:
            query["unit_id"] = {"$in": supervisor_units}
        else:
            raise HTTPException(status_code=403, detail="Supervisor non ha Unit assegnate")
    elif current_user.role == UserRole.SUPER_REFERENTE:
        # Super Referente can export leads assigned to their referenti and agents in their Unit
        super_ref_unit_id = current_user.unit_id
        referenti_ids = current_user.referenti_autorizzati or []
        
        if super_ref_unit_id and referenti_ids:
            # Get all agents under authorized referenti
            agents = await db.users.find({
                "referente_id": {"$in": referenti_ids},
                "is_active": True
            }).to_list(length=None)
            agent_ids = [agent["id"] for agent in agents]
            all_ids = list(set(agent_ids + referenti_ids + [current_user.id]))
            
            query["unit_id"] = super_ref_unit_id
            query["assigned_agent_id"] = {"$in": all_ids}
            logging.info(f"[EXPORT] Super Referente {current_user.username} exporting leads for {len(all_ids)} users in unit {super_ref_unit_id}")
        elif super_ref_unit_id:
            query["unit_id"] = super_ref_unit_id
            query["assigned_agent_id"] = current_user.id
            logging.info(f"[EXPORT] Super Referente {current_user.username} exporting only own leads in unit {super_ref_unit_id}")
        else:
            raise HTTPException(status_code=403, detail="Super Referente non ha Unit assegnata")
    
    # Unit filtering
    if unit_id:
        query["unit_id"] = unit_id
        query["$or"] = [
            {"unit_id": unit_id},
            {"gruppo": unit_id}
        ]
    
    # Apply additional filters - SAME AS GET /leads
    if campagna:
        query["campagna"] = {"$regex": campagna, "$options": "i"}
    if provincia:
        query["provincia"] = {"$regex": provincia, "$options": "i"}
    if status:
        if status == "Nuovo":
            query["$or"] = [
                {"esito": None},
                {"esito": ""},
                {"esito": "Nuovo"},
                {"esito": {"$exists": False}}
            ]
        else:
            query["esito"] = status
    # Filtro date (feb 2026: Europe/Rome → UTC, accetta YYYY-MM-DD o ISO con time)
    if date_from or date_to:
        from helpers import rome_date_to_utc_range
        existing = query.get("created_at") or {}
        if date_from:
            try:
                if "T" in date_from:
                    existing["$gte"] = datetime.fromisoformat(date_from)
                else:
                    start_utc, _ = rome_date_to_utc_range(date_from, current_user.timezone)
                    existing["$gte"] = start_utc
            except ValueError:
                pass
        if date_to:
            try:
                if "T" in date_to:
                    existing["$lte"] = datetime.fromisoformat(date_to)
                else:
                    _, end_utc = rome_date_to_utc_range(date_to, current_user.timezone)
                    existing["$lte"] = end_utc
            except ValueError:
                pass
        if existing:
            query["created_at"] = existing
    
    # NEW: Filter by assigned agent
    if assigned_agent_id:
        if assigned_agent_id == "unassigned":
            query["$or"] = [
                {"assigned_agent_id": None},
                {"assigned_agent_id": {"$exists": False}}
            ]
        else:
            if current_user.role == UserRole.ADMIN or current_user.role == UserRole.REFERENTE:
                query["assigned_agent_id"] = assigned_agent_id
    
    # NEW: Search by name or phone
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"nome": search_regex},
            {"cognome": search_regex},
            {"telefono": search_regex},
            {"email": search_regex}
        ]
    
    # Get custom fields list to add dynamic columns
    custom_fields = await db.custom_fields.find().to_list(length=None)
    
    # Get leads data
    leads = await db.leads.find(query).to_list(length=None)
    
    if not leads:
        raise HTTPException(status_code=404, detail="Nessun lead trovato con i filtri specificati")
    
    # Create Excel file with custom fields
    excel_file_path = await create_excel_report(leads, custom_fields, f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    
    # Return file
    return FileResponse(
        path=excel_file_path,
        filename=f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/analytics/pivot")
async def get_pivot_analytics(
    sub_agenzia_ids: Optional[str] = Query(None),  # Comma-separated IDs
    status_values: Optional[str] = Query(None),  # Comma-separated values
    tipologia_contratto_values: Optional[str] = Query(None),
    segmento_values: Optional[str] = Query(None),
    offerta_ids: Optional[str] = Query(None),
    created_by_ids: Optional[str] = Query(None),
    convergenza: Optional[bool] = Query(None),
    data_da: Optional[str] = Query(None),  # Format: YYYY-MM-DD
    data_a: Optional[str] = Query(None),  # Format: YYYY-MM-DD
    current_user: User = Depends(get_current_user)
):
    """Get pivot analytics with multiple filters and date range"""
    try:
        # Build query
        query = {}
        
        # Role-based access control - Filter by Sub Agenzia, Commessa AND Servizio autorizzati
        if current_user.role == UserRole.ADMIN:
            pass
        elif current_user.role in [UserRole.RESPONSABILE_COMMESSA, UserRole.BACKOFFICE_COMMESSA]:
            if current_user.commesse_autorizzate:
                query["commessa_id"] = {"$in": current_user.commesse_autorizzate}
            else:
                query["_id"] = {"$exists": False}
            if current_user.servizi_autorizzati:
                query["servizio_id"] = {"$in": current_user.servizi_autorizzati}
        elif current_user.role in [UserRole.RESPONSABILE_SUB_AGENZIA, UserRole.BACKOFFICE_SUB_AGENZIA]:
            if current_user.sub_agenzia_id:
                query["sub_agenzia_id"] = current_user.sub_agenzia_id
            else:
                query["_id"] = {"$exists": False}
            if hasattr(current_user, 'commesse_autorizzate') and current_user.commesse_autorizzate:
                query["commessa_id"] = {"$in": current_user.commesse_autorizzate}
            if current_user.servizi_autorizzati:
                query["servizio_id"] = {"$in": current_user.servizi_autorizzati}
        elif current_user.role == UserRole.AREA_MANAGER:
            if hasattr(current_user, 'sub_agenzie_autorizzate') and current_user.sub_agenzie_autorizzate:
                query["sub_agenzia_id"] = {"$in": current_user.sub_agenzie_autorizzate}
            else:
                query["_id"] = {"$exists": False}
            if hasattr(current_user, 'commesse_autorizzate') and current_user.commesse_autorizzate:
                query["commessa_id"] = {"$in": current_user.commesse_autorizzate}
            if current_user.servizi_autorizzati:
                query["servizio_id"] = {"$in": current_user.servizi_autorizzati}
        elif current_user.role == UserRole.RESPONSABILE_PRESIDI:
            if hasattr(current_user, 'sub_agenzie_autorizzate') and current_user.sub_agenzie_autorizzate:
                query["sub_agenzia_id"] = {"$in": current_user.sub_agenzie_autorizzate}
            elif current_user.sub_agenzia_id:
                query["sub_agenzia_id"] = current_user.sub_agenzia_id
            else:
                query["_id"] = {"$exists": False}
            if hasattr(current_user, 'commesse_autorizzate') and current_user.commesse_autorizzate:
                query["commessa_id"] = {"$in": current_user.commesse_autorizzate}
            if current_user.servizi_autorizzati:
                query["servizio_id"] = {"$in": current_user.servizi_autorizzati}
        elif current_user.role in [UserRole.AGENTE_SPECIALIZZATO, UserRole.OPERATORE, UserRole.RESPONSABILE_STORE, UserRole.STORE_ASSIST, UserRole.PROMOTER_PRESIDI]:
            query["$or"] = [
                {"created_by": current_user.id},
                {"assigned_to": current_user.id}
            ]
            if current_user.sub_agenzia_id:
                query["sub_agenzia_id"] = current_user.sub_agenzia_id
            if hasattr(current_user, 'commesse_autorizzate') and current_user.commesse_autorizzate:
                query["commessa_id"] = {"$in": current_user.commesse_autorizzate}
            if current_user.servizi_autorizzati:
                query["servizio_id"] = {"$in": current_user.servizi_autorizzati}
        else:
            query["_id"] = {"$exists": False}
        
        # Apply filters
        if sub_agenzia_ids:
            ids = [id.strip() for id in sub_agenzia_ids.split(",")]
            query["sub_agenzia_id"] = {"$in": ids}
        
        if status_values:
            statuses = [s.strip() for s in status_values.split(",")]
            query["status"] = {"$in": statuses}
        
        if tipologia_contratto_values:
            tipologie = [t.strip() for t in tipologia_contratto_values.split(",")]
            query["tipologia_contratto"] = {"$in": tipologie}
        
        if segmento_values:
            segmenti = [s.strip() for s in segmento_values.split(",")]
            expanded_segmenti = await _expand_segmento_filter_values(segmenti)
            query["segmento"] = {"$in": expanded_segmenti}
        
        if offerta_ids:
            ids = [id.strip() for id in offerta_ids.split(",")]
            query["offerta_id"] = {"$in": ids}
        
        if created_by_ids:
            ids = [id.strip() for id in created_by_ids.split(",")]
            # Use assigned_to instead of created_by to show the assigned user
            query["assigned_to"] = {"$in": ids}
        
        if convergenza is not None:
            query["convergenza"] = convergenza
        
        # Date range filter (feb 2026: input Europe/Rome → UTC, validazione → 400)
        if data_da or data_a:
            from helpers import rome_date_to_utc_range
            date_query = {}
            try:
                if data_da:
                    start_utc, _ = rome_date_to_utc_range(data_da, current_user.timezone)
                    date_query["$gte"] = start_utc
                if data_a:
                    _, end_utc = rome_date_to_utc_range(data_a, current_user.timezone)
                    date_query["$lte"] = end_utc
            except (ValueError, TypeError):
                raise HTTPException(status_code=400, detail="Formato data non valido. Usa YYYY-MM-DD")
            query["created_at"] = date_query
        
        # Get all matching clienti
        clienti = await db.clienti.find(query).to_list(length=None)
        
        # Calculate metrics
        total_clienti = len(clienti)
        
        # Count by sub agenzia
        sub_agenzia_counts = {}
        for cliente in clienti:
            sa_id = cliente.get("sub_agenzia_id", "Non specificato")
            sub_agenzia_counts[sa_id] = sub_agenzia_counts.get(sa_id, 0) + 1
        
        # Count by status
        status_counts = {}
        for cliente in clienti:
            status = cliente.get("status", "Non specificato")
            status_counts[status] = status_counts.get(status, 0) + 1
        
        # Count by tipologia contratto
        tipologia_counts = {}
        for cliente in clienti:
            tipologia = cliente.get("tipologia_contratto", "Non specificato")
            tipologia_counts[tipologia] = tipologia_counts.get(tipologia, 0) + 1
        
        # Count by segmento
        segmento_counts = {}
        for cliente in clienti:
            segmento = cliente.get("segmento", "Non specificato")
            segmento_counts[segmento] = segmento_counts.get(segmento, 0) + 1
        
        # Count by offerta
        offerta_counts = {}
        for cliente in clienti:
            offerta_id = cliente.get("offerta_id", "Non specificato")
            offerta_counts[offerta_id] = offerta_counts.get(offerta_id, 0) + 1
        
        # Count by assigned user (not creator)
        assigned_counts = {}
        for cliente in clienti:
            # Use assigned_to if present, otherwise fall back to created_by
            assigned_id = cliente.get("assigned_to") or cliente.get("created_by", "Non specificato")
            assigned_counts[assigned_id] = assigned_counts.get(assigned_id, 0) + 1
        
        # Count convergenza
        convergenza_counts = {"Si": 0, "No": 0}
        for cliente in clienti:
            if cliente.get("convergenza"):
                convergenza_counts["Si"] += 1
            else:
                convergenza_counts["No"] += 1
        
        # Enrich with names
        enriched_sub_agenzia = {}
        for sa_id, count in sub_agenzia_counts.items():
            if sa_id != "Non specificato":
                sa = await db.sub_agenzie.find_one({"id": sa_id})
                name = sa.get("nome") if sa else sa_id
            else:
                name = "Non specificato"
            enriched_sub_agenzia[name] = count
        
        enriched_offerta = {}
        for off_id, count in offerta_counts.items():
            if off_id != "Non specificato":
                offerta = await db.offerte.find_one({"id": off_id})
                name = offerta.get("nome") if offerta else off_id
            else:
                name = "Non specificato"
            enriched_offerta[name] = count
        
        enriched_assigned = {}
        for assigned_id, count in assigned_counts.items():
            if assigned_id != "Non specificato":
                user = await db.users.find_one({"id": assigned_id})
                name = user.get("username") if user else assigned_id
            else:
                name = "Non specificato"
            enriched_assigned[name] = count
        
        # Enrich segmento with names - INCLUDE ALL SEGMENTS IN THE SYSTEM
        enriched_segmento = {}
        # Fetch all segmenti from dedicated collection
        all_segmenti = await db.segmenti.find({}).to_list(length=None)
        # Map ID -> tipo (name)
        segmenti_map = {seg.get("id"): seg.get("tipo", seg.get("id")) for seg in all_segmenti}
        # Also create a set of valid segment names for direct name matching
        valid_segmento_names = set(seg.get("tipo", "").lower() for seg in all_segmenti if seg.get("tipo"))
        
        # Initialize all segments with 0 count (group by tipo name to avoid duplicates)
        unique_segmento_names = set(segmenti_map.values())
        for seg_name in unique_segmento_names:
            enriched_segmento[seg_name] = 0
        
        # Now add counts from actual clients
        for seg_id, count in segmento_counts.items():
            if seg_id and seg_id not in ["Non specificato", "", None, "None"]:
                # First try to look up as UUID in the map
                if seg_id in segmenti_map:
                    segmento_name = segmenti_map[seg_id]
                # If not found as UUID, check if it's already a valid segment name
                elif seg_id.lower() in valid_segmento_names:
                    # Normalize to proper case (find matching name)
                    segmento_name = seg_id.lower()
                    for name in unique_segmento_names:
                        if name.lower() == seg_id.lower():
                            segmento_name = name
                            break
                else:
                    # Unknown segment, use as-is
                    segmento_name = seg_id
                # Add to existing count (in case multiple IDs map to same name)
                enriched_segmento[segmento_name] = enriched_segmento.get(segmento_name, 0) + count
            else:
                enriched_segmento["Non specificato"] = enriched_segmento.get("Non specificato", 0) + count
        
        # Calculate percentages
        def calc_percentages(counts_dict):
            if total_clienti == 0:
                return {k: 0 for k in counts_dict.keys()}
            return {k: round((v / total_clienti) * 100, 2) for k, v in counts_dict.items()}
        
        # Comparison with previous period (same duration before data_da)
        previous_period_count = 0
        if data_da and data_a:
            try:
                from helpers import rome_date_to_utc_range
                start, _ = rome_date_to_utc_range(data_da, current_user.timezone)
                _, end = rome_date_to_utc_range(data_a, current_user.timezone)
                duration = (end - start).days
                prev_start = start - timedelta(days=duration + 1)
                prev_end = start
                prev_query = query.copy()
                prev_query["created_at"] = {"$gte": prev_start, "$lt": prev_end}
                previous_period_count = await db.clienti.count_documents(prev_query)
            except (ValueError, TypeError):
                pass  # already validated above; difensivo
        
        # Calculate trend
        trend = None
        if previous_period_count > 0:
            trend = round(((total_clienti - previous_period_count) / previous_period_count) * 100, 2)
        
        return {
            "total_clienti": total_clienti,
            "previous_period_count": previous_period_count,
            "trend_percentage": trend,
            "breakdown": {
                "sub_agenzia": {
                    "counts": enriched_sub_agenzia,
                    "percentages": calc_percentages(enriched_sub_agenzia)
                },
                "status": {
                    "counts": status_counts,
                    "percentages": calc_percentages(status_counts)
                },
                "tipologia_contratto": {
                    "counts": tipologia_counts,
                    "percentages": calc_percentages(tipologia_counts)
                },
                "segmento": {
                    "counts": enriched_segmento,
                    "percentages": calc_percentages(enriched_segmento)
                },
                "offerta": {
                    "counts": enriched_offerta,
                    "percentages": calc_percentages(enriched_offerta)
                },
                "assigned_to": {
                    "counts": enriched_assigned,
                    "percentages": calc_percentages(enriched_assigned)
                },
                "convergenza": {
                    "counts": convergenza_counts,
                    "percentages": calc_percentages(convergenza_counts)
                }
            }
        }
        
    except HTTPException:
        # NEW (feb 2026): lascia passare le HTTPException già strutturate (es. 400 su date invalide)
        raise
    except Exception as e:
        logging.error(f"Error in pivot analytics: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Errore analytics pivot: {str(e)}")


@router.get("/analytics/sub-agenzie")
async def get_sub_agenzie_analytics(
    data_da: Optional[str] = Query(None),
    data_a: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """Get analytics for each sub agenzia"""
    try:
        # Build base query with date filter
        base_query = {}
        
        if data_da or data_a:
            date_query = {}
            if data_da:
                date_query["$gte"] = datetime.strptime(data_da, "%Y-%m-%d")
            if data_a:
                end_date = datetime.strptime(data_a, "%Y-%m-%d") + timedelta(days=1)
                date_query["$lt"] = end_date
            base_query["created_at"] = date_query
        
        # Get all sub agenzie based on user role
        if current_user.role == UserRole.ADMIN:
            sub_agenzie = await db.sub_agenzie.find({}).to_list(length=None)
        elif current_user.role in [UserRole.RESPONSABILE_SUB_AGENZIA, UserRole.BACKOFFICE_SUB_AGENZIA]:
            if current_user.sub_agenzia_id:
                sub_agenzie = await db.sub_agenzie.find({"id": current_user.sub_agenzia_id}).to_list(length=1)
            else:
                sub_agenzie = []
        elif current_user.role in [UserRole.RESPONSABILE_COMMESSA, UserRole.BACKOFFICE_COMMESSA, UserRole.AREA_MANAGER]:
            # For these roles, get sub agenzie from their authorized commesse
            if current_user.commesse_autorizzate:
                sub_agenzie_query = {
                    "commesse_autorizzate": {"$in": current_user.commesse_autorizzate}
                }
                # Filter by authorized services
                if current_user.servizi_autorizzati:
                    sub_agenzie_query["servizi_autorizzati"] = {"$in": current_user.servizi_autorizzati}
                
                sub_agenzie = await db.sub_agenzie.find(sub_agenzie_query).to_list(length=None)
            else:
                sub_agenzie = []
        else:
            # No access for other roles
            sub_agenzie = []
        
        result = []
        
        for sub_agenzia in sub_agenzie:
            sa_id = sub_agenzia["id"]
            
            # Query for this sub agenzia
            query = base_query.copy()
            query["sub_agenzia_id"] = sa_id
            
            # Apply role-based filters - Sub Agenzia, Commessa AND Servizio
            if current_user.role in [UserRole.RESPONSABILE_COMMESSA, UserRole.BACKOFFICE_COMMESSA, UserRole.AREA_MANAGER]:
                if current_user.commesse_autorizzate:
                    query["commessa_id"] = {"$in": current_user.commesse_autorizzate}
                if current_user.servizi_autorizzati:
                    query["servizio_id"] = {"$in": current_user.servizi_autorizzati}
            elif current_user.role in [UserRole.RESPONSABILE_SUB_AGENZIA, UserRole.BACKOFFICE_SUB_AGENZIA]:
                if hasattr(current_user, 'commesse_autorizzate') and current_user.commesse_autorizzate:
                    query["commessa_id"] = {"$in": current_user.commesse_autorizzate}
                if current_user.servizi_autorizzati:
                    query["servizio_id"] = {"$in": current_user.servizi_autorizzati}
            
            clienti = await db.clienti.find(query).to_list(length=None)
            
            # Total clienti
            total = len(clienti)
            
            # Clienti by status
            status_breakdown = {}
            for cliente in clienti:
                status = cliente.get("status", "Non specificato")
                status_breakdown[status] = status_breakdown.get(status, 0) + 1
            
            # Clienti by tipologia contratto
            tipologia_breakdown = {}
            for cliente in clienti:
                tipologia = cliente.get("tipologia_contratto", "Non specificato")
                tipologia_breakdown[tipologia] = tipologia_breakdown.get(tipologia, 0) + 1
            
            # Top assigned users (not creators)
            assigned_counts = {}
            for cliente in clienti:
                # Use assigned_to if present, otherwise fall back to created_by
                assigned_id = cliente.get("assigned_to") or cliente.get("created_by")
                if assigned_id:
                    assigned_counts[assigned_id] = assigned_counts.get(assigned_id, 0) + 1
            
            # Enrich assigned user names
            top_assigned = []
            for assigned_id, count in sorted(assigned_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
                user = await db.users.find_one({"id": assigned_id})
                top_assigned.append({
                    "name": user.get("username") if user else assigned_id,
                    "count": count
                })
            
            # Performance over time (last 30 days or date range)
            timeline = []
            if data_da and data_a:
                start = datetime.strptime(data_da, "%Y-%m-%d")
                end = datetime.strptime(data_a, "%Y-%m-%d")
                days = (end - start).days + 1
                
                for i in range(days):
                    day = start + timedelta(days=i)
                    day_end = day + timedelta(days=1)
                    
                    day_query = query.copy()
                    day_query["created_at"] = {"$gte": day, "$lt": day_end}
                    
                    count = await db.clienti.count_documents(day_query)
                    timeline.append({
                        "date": day.strftime("%Y-%m-%d"),
                        "count": count
                    })
            
            result.append({
                "sub_agenzia_id": sa_id,
                "sub_agenzia_name": sub_agenzia.get("nome"),
                "total_clienti": total,
                "status_breakdown": status_breakdown,
                "tipologia_breakdown": tipologia_breakdown,
                "top_assigned": top_assigned,
                "timeline": timeline
            })
        
        return result
        
    except Exception as e:
        logging.error(f"Error in sub agenzie analytics: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Errore analytics sub agenzie: {str(e)}")


@router.get("/analytics/pivot/export")
async def export_pivot_analytics(
    sub_agenzia_ids: Optional[str] = Query(None),
    status_values: Optional[str] = Query(None),
    tipologia_contratto_values: Optional[str] = Query(None),
    segmento_values: Optional[str] = Query(None),
    offerta_ids: Optional[str] = Query(None),
    created_by_ids: Optional[str] = Query(None),
    convergenza: Optional[bool] = Query(None),
    data_da: Optional[str] = Query(None),
    data_a: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """Export pivot analytics to Excel"""
    try:
        # Get pivot data using the same logic
        pivot_data = await get_pivot_analytics(
            sub_agenzia_ids=sub_agenzia_ids,
            status_values=status_values,
            tipologia_contratto_values=tipologia_contratto_values,
            segmento_values=segmento_values,
            offerta_ids=offerta_ids,
            created_by_ids=created_by_ids,
            convergenza=convergenza,
            data_da=data_da,
            data_a=data_a,
            current_user=current_user
        )
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pivot Analytics"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Summary section
        ws.cell(1, 1, "RIEPILOGO").font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(2, 1, "Totale Clienti")
        ws.cell(2, 2, pivot_data["total_clienti"])
        ws.cell(3, 1, "Periodo Precedente")
        ws.cell(3, 2, pivot_data["previous_period_count"])
        ws.cell(4, 1, "Trend %")
        ws.cell(4, 2, pivot_data["trend_percentage"])
        
        # Breakdown sections
        row = 6
        for category, data in pivot_data["breakdown"].items():
            ws.cell(row, 1, category.upper().replace("_", " ")).font = header_font
            ws.cell(row, 1).fill = header_fill
            ws.cell(row, 2, "Conteggio").font = header_font
            ws.cell(row, 2).fill = header_fill
            ws.cell(row, 3, "Percentuale").font = header_font
            ws.cell(row, 3).fill = header_fill
            row += 1
            
            for key, count in data["counts"].items():
                ws.cell(row, 1, key)
                ws.cell(row, 2, count)
                ws.cell(row, 3, f"{data['percentages'].get(key, 0)}%")
                row += 1
            
            row += 1
        
        # Save file
        filename = f"pivot_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"/tmp/{filename}"
        wb.save(filepath)
        
        return FileResponse(
            path=filepath,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=filename
        )
        
    except Exception as e:
        logging.error(f"Error in pivot export: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Errore export pivot: {str(e)}")


@router.get("/analytics/pivot/export-clienti")
async def export_pivot_clienti_list(
    sub_agenzia_ids: Optional[str] = Query(None),
    status_values: Optional[str] = Query(None),
    tipologia_contratto_values: Optional[str] = Query(None),
    segmento_values: Optional[str] = Query(None),
    offerta_ids: Optional[str] = Query(None),
    created_by_ids: Optional[str] = Query(None),
    convergenza: Optional[bool] = Query(None),
    data_da: Optional[str] = Query(None),
    data_a: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """Export lista completa clienti filtrati (stesso criterio del pivot) in formato Excel"""
    try:
        from datetime import datetime, timezone
        
        # Build query based on filters
        query = {}
        
        # Role-based access control
        if current_user.role == UserRole.ADMIN:
            pass  # Admin can see all
        elif current_user.role in [UserRole.RESPONSABILE_COMMESSA, UserRole.BACKOFFICE_COMMESSA]:
            if current_user.commesse_autorizzate:
                query["commessa_id"] = {"$in": current_user.commesse_autorizzate}
            else:
                query["_id"] = {"$exists": False}
        elif current_user.role in [UserRole.RESPONSABILE_SUB_AGENZIA, UserRole.BACKOFFICE_SUB_AGENZIA]:
            if current_user.sub_agenzia_id:
                query["sub_agenzia_id"] = current_user.sub_agenzia_id
            else:
                query["_id"] = {"$exists": False}
        elif current_user.role == UserRole.RESPONSABILE_PRESIDI:
            if hasattr(current_user, 'sub_agenzie_autorizzate') and current_user.sub_agenzie_autorizzate:
                query["sub_agenzia_id"] = {"$in": current_user.sub_agenzie_autorizzate}
            elif current_user.sub_agenzia_id:
                query["sub_agenzia_id"] = current_user.sub_agenzia_id
            else:
                query["_id"] = {"$exists": False}
        elif current_user.role == UserRole.AREA_MANAGER:
            if hasattr(current_user, 'sub_agenzie_autorizzate') and current_user.sub_agenzie_autorizzate:
                query["sub_agenzia_id"] = {"$in": current_user.sub_agenzie_autorizzate}
            else:
                query["_id"] = {"$exists": False}
        else:
            query["_id"] = {"$exists": False}
        
        # Apply pivot filters
        if sub_agenzia_ids:
            sa_list = [s.strip() for s in sub_agenzia_ids.split(',') if s.strip()]
            if sa_list:
                query["sub_agenzia_id"] = {"$in": sa_list}
        
        if status_values:
            status_list = [s.strip() for s in status_values.split(',') if s.strip()]
            if status_list:
                query["status"] = {"$in": status_list}
        
        if tipologia_contratto_values:
            tipologia_list = [t.strip() for t in tipologia_contratto_values.split(',') if t.strip()]
            if tipologia_list:
                query["tipologia_contratto"] = {"$in": tipologia_list}
        
        if segmento_values:
            segmento_list = [s.strip() for s in segmento_values.split(',') if s.strip()]
            if segmento_list:
                expanded_segmenti = await _expand_segmento_filter_values(segmento_list)
                query["segmento"] = {"$in": expanded_segmenti}
        
        if offerta_ids:
            offerta_list = [o.strip() for o in offerta_ids.split(',') if o.strip()]
            if offerta_list:
                query["offerta_id"] = {"$in": offerta_list}
        
        if created_by_ids:
            created_list = [c.strip() for c in created_by_ids.split(',') if c.strip()]
            if created_list:
                query["assigned_to"] = {"$in": created_list}
        
        if convergenza is not None:
            query["convergenza"] = convergenza
        
        # Date filters
        if data_da or data_a:
            date_query = {}
            if data_da:
                date_query["$gte"] = datetime.strptime(data_da, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            if data_a:
                date_query["$lte"] = datetime.strptime(data_a, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
            if date_query:
                query["created_at"] = date_query
        
        # Fetch clienti
        clienti = await db.clienti.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)
        
        # Fetch lookup data
        sub_agenzie = await db.sub_agenzie.find({}, {"_id": 0}).to_list(None)
        sub_agenzie_map = {sa["id"]: sa["nome"] for sa in sub_agenzie}
        
        users_list = await db.users.find({}, {"_id": 0, "id": 1, "username": 1}).to_list(None)
        users_map = {u["id"]: u["username"] for u in users_list}
        
        commesse_list = await db.commesse.find({}, {"_id": 0, "id": 1, "nome": 1}).to_list(None)
        commesse_map = {c["id"]: c["nome"] for c in commesse_list}
        
        servizi_list = await db.servizi.find({}, {"_id": 0, "id": 1, "nome": 1}).to_list(None)
        servizi_map = {s["id"]: s["nome"] for s in servizi_list}
        
        # Fetch segmenti for name lookup
        segmenti_list = await db.segmenti.find({}, {"_id": 0, "id": 1, "nome": 1}).to_list(None)
        segmenti_map = {s["id"]: s["nome"] for s in segmenti_list}
        
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista Clienti"
        
        # Headers
        headers = [
            "ID", "Nome", "Cognome", "Codice Fiscale", "Partita IVA",
            "Telefono", "Email", "Indirizzo", "Comune", "Provincia", "CAP",
            "Sub Agenzia", "Commessa", "Servizio", "Status", "Tipologia Contratto",
            "Segmento", "Convergenza", "Assegnato a", "Data Creazione", "Note"
        ]
        
        # Fetch active custom fields for dynamic columns
        try:
            raw_cf = await db.cliente_custom_fields.find({"active": True}, {"_id": 0}).to_list(length=None)
            seen_cf = {}
            for f in raw_cf:
                nm = f.get("name")
                if nm and nm not in seen_cf:
                    seen_cf[nm] = f.get("label") or nm
            export_custom_fields = [{"name": k, "label": v} for k, v in sorted(seen_cf.items(), key=lambda x: x[1].lower())]
        except Exception as _cf_err:
            logging.warning(f"Could not fetch custom fields for pivot export: {_cf_err}")
            export_custom_fields = []
        
        for cf in export_custom_fields:
            headers.append(f"[Custom] {cf['label']}")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row_num, cliente in enumerate(clienti, 2):
            ws.cell(row_num, 1, cliente.get("id", ""))
            ws.cell(row_num, 2, cliente.get("nome", ""))
            ws.cell(row_num, 3, cliente.get("cognome", ""))
            ws.cell(row_num, 4, cliente.get("codice_fiscale", ""))
            ws.cell(row_num, 5, cliente.get("partita_iva", ""))
            ws.cell(row_num, 6, cliente.get("telefono", ""))
            ws.cell(row_num, 7, cliente.get("email", ""))
            ws.cell(row_num, 8, cliente.get("indirizzo", ""))
            ws.cell(row_num, 9, cliente.get("comune", ""))
            ws.cell(row_num, 10, cliente.get("provincia", ""))
            ws.cell(row_num, 11, cliente.get("cap", ""))
            ws.cell(row_num, 12, sub_agenzie_map.get(cliente.get("sub_agenzia_id", ""), ""))
            ws.cell(row_num, 13, commesse_map.get(cliente.get("commessa_id", ""), ""))
            ws.cell(row_num, 14, servizi_map.get(cliente.get("servizio_id", ""), ""))
            ws.cell(row_num, 15, cliente.get("status", ""))
            ws.cell(row_num, 16, cliente.get("tipologia_contratto", ""))
            # Segmento: lookup name from ID
            segmento_id = cliente.get("segmento", "")
            ws.cell(row_num, 17, segmenti_map.get(segmento_id, segmento_id))  # Fallback to ID if not found
            ws.cell(row_num, 18, "Sì" if cliente.get("convergenza") else "No")
            ws.cell(row_num, 19, users_map.get(cliente.get("assigned_to", ""), ""))
            
            created_at = cliente.get("created_at")
            if created_at:
                if isinstance(created_at, str):
                    ws.cell(row_num, 20, created_at)
                else:
                    ws.cell(row_num, 20, created_at.strftime("%d/%m/%Y %H:%M"))
            else:
                ws.cell(row_num, 20, "")
            
            ws.cell(row_num, 21, cliente.get("note", ""))
            
            # Custom fields columns
            cf_col = 22
            dati_agg = cliente.get("dati_aggiuntivi") or {}
            for cf in export_custom_fields:
                v = dati_agg.get(cf["name"], "")
                if isinstance(v, list):
                    v = ", ".join(str(x) for x in v)
                elif isinstance(v, bool):
                    v = "Sì" if v else "No"
                ws.cell(row_num, cf_col, v if v is not None else "")
                cf_col += 1
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save
        filename = f"clienti_pivot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = f"/tmp/{filename}"
        wb.save(filepath)
        
        return FileResponse(
            path=filepath,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=filename
        )
        
    except Exception as e:
        logging.error(f"Error in pivot clienti export: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Errore export clienti: {str(e)}")


